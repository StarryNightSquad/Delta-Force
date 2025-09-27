print("本程序由繁星攻略组制作")

import sys
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
import msvcrt
import os

# 全局调试标志
DEBUG_MODE = False

def debug_print(*args, **kwargs):
    """只在调试模式下打印信息"""
    if DEBUG_MODE:
        print("[DEBUG]", *args, **kwargs)

def get_decimal_input(prompt, min_val, max_val, decimal_places):
    """获取精确的小数输入"""
    while True:
        try:
            value = input(prompt)
            decimal_value = Decimal(value)
            quantized = decimal_value.quantize(
                Decimal('1.' + '0' * decimal_places), 
                rounding=ROUND_HALF_UP
            )
            float_val = float(quantized)
            if min_val <= float_val <= max_val:
                return quantized
            else:
                print(f"输入错误，范围{min_val}到{max_val}，请重新输入。")
        except Exception:
            print("输入错误，请输入有效的数字。")

def get_int_input(prompt, min_val, max_val):
    """获取整数输入"""
    while True:
        try:
            value = int(input(prompt))
            if min_val <= value <= max_val:
                return value
            else:
                print(f"输入错误，范围{min_val}到{max_val}，请重新输入。")
        except ValueError:
            print("输入错误，请输入整数。")

def wait_for_key(continue_msg="按 Enter 开始新一轮计算，按 ESC 结束程序"):
    """等待用户按键，返回True继续，False退出"""
    print(f"\n{continue_msg}")
    while True:
        if msvcrt.kbhit():
            key = msvcrt.getch()
            if key == b'\r':  # Enter键
                return True
            elif key == b'\x1b':  # ESC键
                return False

def standardize_caliber(caliber):
    """标准化口径表示，确保匹配"""
    if caliber is None:
        return ""
    # 去除空格和点号，统一大小写
    return caliber.replace(" ", "").replace(".", "").lower().strip()

def load_weapons_data():
    """加载武器数据（修复版本）"""
    try:
        wb = openpyxl.load_workbook('S6夺金武器.xlsx', data_only=True)
        ws = wb['夺金模式']
        
        weapons = []
        current_category = ""
        
        for row in range(3, ws.max_row + 1):
            # 读取关键单元格
            category_cell = ws.cell(row=row, column=1).value
            weapon_name = ws.cell(row=row, column=2).value
            
            # 更新当前武器类别（如果有）
            if category_cell and str(category_cell).strip():
                current_category = str(category_cell).strip()
            
            # 跳过空行（没有武器名称的行）
            if not weapon_name or not str(weapon_name).strip():
                continue
            
            # 标准化口径
            caliber = ws.cell(row=row, column=3).value
            std_caliber = standardize_caliber(caliber) if caliber else ""
            
            # 创建武器数据字典
            weapon_data = {
                'category': current_category,
                'name': str(weapon_name).strip(),
                'caliber': std_caliber,
                'raw_caliber': caliber,  # 保留原始口径用于显示
                'fire_mode': ws.cell(row=row, column=5).value or 0,
                'trigger_delay': ws.cell(row=row, column=6).value or 0,
                'fire_rate': ws.cell(row=row, column=7).value or 0,
                'muzzle_velocity': ws.cell(row=row, column=8).value or 0,
                'base_damage': ws.cell(row=row, column=9).value or 0,
                'armor_damage': ws.cell(row=row, column=11).value or 0,
                # 部位倍率
                'head_multiplier': ws.cell(row=row, column=13).value or 1.0,
                'chest_multiplier': ws.cell(row=row, column=14).value or 1.0,
                'abdomen_multiplier': ws.cell(row=row, column=15).value or 1.0,
                'upper_arm_multiplier': ws.cell(row=row, column=16).value or 1.0,
                'lower_arm_multiplier': ws.cell(row=row, column=17).value or 1.0,
                'thigh_multiplier': ws.cell(row=row, column=18).value or 1.0,
                'calf_multiplier': ws.cell(row=row, column=19).value or 1.0,
                'decay_distances': [],
                'decay_factors': []
            }
            
            # 读取衰减数据（最多4组）
            decay_cols = [(20, 21), (22, 23), (24, 25), (26, 27)]  # T/U, V/W, X/Y, Z/AA列
            
            for dist_col, factor_col in decay_cols:
                dist = ws.cell(row=row, column=dist_col).value
                factor = ws.cell(row=row, column=factor_col).value
                
                # 处理特殊值
                if dist in (None, "", "/", "//", "N/A"):
                    continue
                    
                try:
                    # 尝试转换为数值
                    dist_val = float(dist)
                    factor_val = float(factor)
                    weapon_data['decay_distances'].append(dist_val)
                    weapon_data['decay_factors'].append(factor_val)
                except (ValueError, TypeError):
                    continue
            
            weapons.append(weapon_data)
        
        debug_print(f"成功加载 {len(weapons)} 种武器")
        return weapons
    
    except Exception as e:
        print(f"\n错误: 无法加载武器数据 - {e}")
        return []

def load_bullets_data():
    """加载子弹数据（修复版本）"""
    try:
        wb = openpyxl.load_workbook('S6子弹数据.xlsx', data_only=True)
        ws = wb['子弹数据']
        
        bullets = []
        current_caliber = ""
        
        for row in range(3, ws.max_row + 1):
            # 读取关键单元格
            caliber_cell = ws.cell(row=row, column=1).value
            bullet_name = ws.cell(row=row, column=2).value
            
            # 更新当前口径（如果有）
            if caliber_cell and str(caliber_cell).strip():
                current_caliber = standardize_caliber(caliber_cell)
            
            # 跳过空行（没有子弹名称的行）
            if not bullet_name or not str(bullet_name).strip():
                continue
                
            # 处理可能的空值
            pellet_count = ws.cell(row=row, column=3).value or 1
            penetration_level = ws.cell(row=row, column=4).value or 0
            base_damage_multiplier = ws.cell(row=row, column=6).value or 1.0
            base_armor_multiplier = ws.cell(row=row, column=7).value or 1.0
            
            # 读取新的穿透倍率数据（N列和O列）
            same_level_penetration = ws.cell(row=row, column=14).value or 0.5  # N列：同级穿透倍率
            higher_level_penetration = ws.cell(row=row, column=15).value or 0.75  # O列：越级穿透倍率
            
            bullet_data = {
                'caliber': current_caliber,
                'name': str(bullet_name).strip(),
                'pellet_count': pellet_count,
                'penetration_level': penetration_level,
                'base_damage_multiplier': base_damage_multiplier,
                'base_armor_multiplier': base_armor_multiplier,
                'same_level_penetration': same_level_penetration,  # 新增：同级穿透倍率
                'higher_level_penetration': higher_level_penetration,  # 新增：越级穿透倍率
                'armor_decay_factors': []
            }
            
            # 读取护甲衰减倍率 (1-6级)
            for col in range(7, 13):  # G到L列
                factor = ws.cell(row=row, column=col).value
                if factor is None:
                    factor = 0.0
                try:
                    bullet_data['armor_decay_factors'].append(float(factor))
                except (ValueError, TypeError):
                    bullet_data['armor_decay_factors'].append(0.0)
            
            bullets.append(bullet_data)
        
        debug_print(f"成功加载 {len(bullets)} 种子弹")
        return bullets
    
    except Exception as e:
        print(f"\n错误: 无法加载子弹数据 - {e}")
        return []

def calculate_weapon_decay(distance, weapon):
    """计算武器衰减倍率"""
    if not weapon['decay_distances']:
        debug_print(f"武器 {weapon['name']} 无衰减数据，使用默认倍率 1.0")
        return Decimal('1.0')
    
    # 如果距离小于第一个衰减距离，无衰减
    if distance <= weapon['decay_distances'][0]:
        debug_print(f"距离 {distance} 小于第一个衰减距离 {weapon['decay_distances'][0]}，无衰减")
        return Decimal('1.0')
    
    # 检查后续衰减距离
    for i in range(1, len(weapon['decay_distances'])):
        if distance <= weapon['decay_distances'][i]:
            decay_factor = Decimal(str(weapon['decay_factors'][i-1]))
            debug_print(f"距离 {distance} 在衰减区间 {weapon['decay_distances'][i-1]}-{weapon['decay_distances'][i]}，衰减倍率: {decay_factor}")
            return decay_factor
    
    # 如果超过所有衰减距离，使用最后一个衰减倍率
    decay_factor = Decimal(str(weapon['decay_factors'][-1]))
    debug_print(f"距离 {distance} 超过所有衰减距离，使用最后衰减倍率: {decay_factor}")
    return decay_factor

def load_armor_data():
    """加载护甲和头盔数据（修复版本）"""
    try:
        wb = openpyxl.load_workbook('S6护甲数据.xlsx', data_only=True)
        ws = wb.active
        
        armors = []
        helmets = []
        current_section = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            
            # 检测章节开始标记
            if cell_value == "护甲":
                current_section = "armor"
            elif cell_value == "头盔":
                current_section = "helmet"
            
            # 读取数据列
            name = str(cell_value).strip() if cell_value else ""
            level = ws.cell(row=row, column=2).value
            armor_type = ws.cell(row=row, column=3).value
            max_durability = ws.cell(row=row, column=7).value  # G列: 初始上限
            
            # 跳过空行和标记行（没有有效数据的行）
            if not name or name in ["护甲", "头盔"] or level is None or max_durability is None:
                continue
                
            try:
                level = int(level)
                max_durability = Decimal(str(max_durability))
            except (ValueError, TypeError):
                continue
                
            item_data = {
                'name': name,
                'level': level,
                'max_durability': max_durability
            }
            
            if current_section == "armor":
                # 护甲类型映射
                if armor_type == '半甲':
                    item_data['armor_type'] = 1
                elif armor_type == '全甲':
                    item_data['armor_type'] = 2
                elif armor_type == '重甲':
                    item_data['armor_type'] = 3
                else:
                    continue  # 跳过无效类型
                armors.append(item_data)
                
            elif current_section == "helmet":
                helmets.append(item_data)
                
        debug_print(f"成功加载 {len(armors)} 种护甲和 {len(helmets)} 种头盔")
        return armors, helmets
    
    except Exception as e:
        print(f"\n错误: 无法加载护甲数据 - {e}")
        return [], []

def select_protection(items, item_type):
    """选择防护装备并输入耐久"""
    print(f"\n=== 选择{item_type} ===")
    print("0. 无")
    
    # 按等级分组
    levels = {}
    for item in items:
        level = item['level']
        if level not in levels:
            levels[level] = []
        levels[level].append(item)
    
    # 显示等级选项
    sorted_levels = sorted(levels.keys())
    for i, level in enumerate(sorted_levels, 1):
        print(f"{i}. {level}级{item_type}")
    
    # 获取用户选择，包括0（无）选项
    max_choice = len(sorted_levels)
    level_choice = get_int_input(f"请选择{item_type}等级 (0-{max_choice}): ", 0, max_choice)
    
    # 处理"无"选项
    if level_choice == 0:
        debug_print(f"用户选择无{item_type}")
        return None, Decimal('0.0')
    
    selected_level = sorted_levels[level_choice - 1]
    
    # 显示该等级下的装备
    level_items = levels[selected_level]
    print(f"\n{selected_level}级{item_type}列表:")
    for i, item in enumerate(level_items, 1):
        print(f"{i}. {item['name']} (最大耐久: {item['max_durability']})")
    
    item_choice = get_int_input(f"请选择{item_type}: ", 1, len(level_items))
    selected_item = level_items[item_choice - 1]
    
    # 输入当前耐久
    max_durability = float(selected_item['max_durability'])
    durability = get_decimal_input(
        f"请输入当前{item_type}耐久 (0.0-{max_durability}): ",
        0.0, max_durability, 1
    )
    
    debug_print(f"选择{item_type}: {selected_item['name']}, 等级: {selected_item['level']}, 耐久: {durability}")
    return selected_item, durability

def calculate_penetration_multiplier(penetration_level, protector_level, bullet):
    """计算穿透倍率（使用子弹数据中的穿透倍率）"""
    diff = penetration_level - protector_level
    
    if diff < 0:
        # 子弹穿透等级低于护甲等级，无法穿透
        return Decimal('0.0')
    elif diff == 0:
        # 同级穿透
        same_level_pen = Decimal(str(bullet.get('same_level_penetration', 0.5)))
        debug_print(f"同级穿透，使用倍率: {same_level_pen}")
        return same_level_pen
    elif diff == 1:
        # 越一级穿透
        higher_level_pen = Decimal(str(bullet.get('higher_level_penetration', 0.75)))
        debug_print(f"越一级穿透，使用倍率: {higher_level_pen}")
        return higher_level_pen
    else:
        # 越多级穿透，完全穿透
        debug_print(f"越{diff}级穿透，完全穿透")
        return Decimal('1.0')

def run_simulation():
    """运行一次完整的伤害模拟"""
    # 初始化参数
    print("=== 通用武器伤害模拟器 ===")
    
    # 加载护甲数据
    print("正在加载护甲数据...")
    armors, helmets = load_armor_data()
    
    # 选择头盔
    helmet_level = 0
    helmet_durability = Decimal('0.0')
    
    if helmets:
        selected_helmet, helmet_durability = select_protection(helmets, "头盔")
        if selected_helmet:
            helmet_level = selected_helmet['level']
            print(f"已选择头盔: {selected_helmet['name']} (等级{helmet_level})")
        else:
            print("已选择：无头盔")
    else:
        print("未找到头盔数据，将使用无头盔设置")
    
    # 选择护甲
    armor_level = 0
    armor_durability = Decimal('0.0')
    armor_type_value = 0
    
    if armors:
        selected_armor, armor_durability = select_protection(armors, "护甲")
        if selected_armor:
            armor_level = selected_armor['level']
            armor_type_value = selected_armor['armor_type']
            print(f"已选择护甲: {selected_armor['name']} (等级{armor_level}, 类型{armor_type_value})")
        else:
            print("已选择：无护甲")
    else:
        print("未找到护甲数据，将使用无护甲设置")
    
    # 设置保护的身体部位
    protected_areas = {
        1: ['胸部', '腹部'],
        2: ['胸部', '腹部', '下腹部'],
        3: ['胸部', '腹部', '下腹部', '大臂']
    }.get(armor_type_value, [])  # 如果没有护甲，返回空列表
    
    debug_print(f"护甲类型: {armor_type_value}, 保护部位: {protected_areas}")
    
    # 加载武器和子弹数据
    print("正在加载武器数据...")
    weapons = load_weapons_data()
    if not weapons:
        print("无法加载武器数据，程序退出")
        return False
    
    print("正在加载子弹数据...")
    bullets = load_bullets_data()
    if not bullets:
        print("无法加载子弹数据，程序退出")
        return False
    
    # 选择武器
    print("\n=== 选择武器 ===")
    print("（点射武器与多弹丸武器不适用本程序）")
    
    # 按类别分组武器
    categories = {}
    for weapon in weapons:
        if weapon['category'] not in categories:
            categories[weapon['category']] = []
        categories[weapon['category']].append(weapon)
    
    # 显示武器类别
    print("\n请选择武器类别：")
    category_list = list(categories.keys())
    for i, category in enumerate(category_list, 1):
        print(f"{i}. {category}")
    
    category_choice = get_int_input("输入类别编号：", 1, len(category_list))
    selected_category = category_list[category_choice - 1]
    
    # 显示选定类别中的武器
    print(f"\n{selected_category}武器列表：")
    category_weapons = categories[selected_category]
    for i, weapon in enumerate(category_weapons, 1):
        print(f"{i}. {weapon['name']} ({weapon['raw_caliber']})")
    
    weapon_choice = get_int_input("输入武器编号：", 1, len(category_weapons))
    selected_weapon = category_weapons[weapon_choice - 1]
    weapon_caliber = selected_weapon['caliber']
    
    debug_print(f"选择武器: {selected_weapon['name']}, 口径: {weapon_caliber}")
    
    # 霰弹枪警告
    if selected_weapon['category'] == "霰弹枪":
        print("警告：霰弹枪伤害计算可能不准确（多弹丸特性）")
    
    # 根据武器口径过滤子弹
    caliber_bullets = [b for b in bullets if b['caliber'] == weapon_caliber]
    
    if not caliber_bullets:
        print(f"\n警告：没有找到匹配 {weapon_caliber} 口径的子弹！")
        print(f"武器口径: {selected_weapon['raw_caliber']}")
        print("可用的子弹口径:")
        unique_calibers = set(b['caliber'] for b in bullets)
        for cal in unique_calibers:
            print(f" - {cal}")
        print("程序将退出。")
        return False
    
    # 选择子弹
    print(f"\n=== 选择 {selected_weapon['raw_caliber']} 子弹 ===")
    for i, bullet in enumerate(caliber_bullets, 1):
        pen_level = bullet['penetration_level']
        dmg_mult = bullet['base_damage_multiplier']
        same_pen = bullet['same_level_penetration']
        higher_pen = bullet['higher_level_penetration']
        print(f"{i}. {bullet['name']} (穿透:{pen_level}, 伤害:{dmg_mult}, 同级穿透:{same_pen}, 越级穿透:{higher_pen})")
    
    bullet_choice = get_int_input("输入子弹编号：", 1, len(caliber_bullets))
    selected_bullet = caliber_bullets[bullet_choice - 1]
    
    debug_print(f"选择子弹: {selected_bullet['name']}, 穿透等级: {selected_bullet['penetration_level']}, 伤害倍率: {selected_bullet['base_damage_multiplier']}")
    debug_print(f"同级穿透倍率: {selected_bullet['same_level_penetration']}, 越级穿透倍率: {selected_bullet['higher_level_penetration']}")
    
    # 输入目标距离
    distance = get_decimal_input("\n请输入目标距离（0-400米）：", 0.0, 400.0, 1)
    
    # 计算武器衰减倍率
    weapon_decay_multiplier = calculate_weapon_decay(float(distance), selected_weapon)
    
    # 初始化头盔和护甲衰减倍率
    helmet_decay_multiplier = Decimal('0.0')
    armor_decay_multiplier = Decimal('0.0')
    
    # 计算头盔衰减倍率 (根据头盔等级)
    if helmet_level > 0 and helmet_level <= 6:
        helmet_decay_multiplier = Decimal(str(
            selected_bullet['armor_decay_factors'][helmet_level - 1]
        ))
        debug_print(f"头盔等级 {helmet_level}, 头盔衰减倍率: {helmet_decay_multiplier}")
    
    # 计算护甲衰减倍率 (根据护甲等级)
    if armor_level > 0 and armor_level <= 6:
        armor_decay_multiplier = Decimal(str(
            selected_bullet['armor_decay_factors'][armor_level - 1]
        ))
        debug_print(f"护甲等级 {armor_level}, 护甲衰减倍率: {armor_decay_multiplier}")
    
    # 设置武器参数
    weapon_damage = selected_weapon['base_damage']
    weapon_armor_damage = selected_weapon['armor_damage']
    fire_rate = selected_weapon['fire_rate']
    fire_mode = selected_weapon['fire_mode']
    trigger_delay = selected_weapon['trigger_delay']
    
    debug_print(f"武器基础伤害: {weapon_damage}, 武器护甲伤害: {weapon_armor_damage}")
    debug_print(f"射速: {fire_rate}, 射击模式: {fire_mode}, 扳机延迟: {trigger_delay}")
    
    # 设置子弹参数
    penetration_level = selected_bullet['penetration_level']
    base_damage_multiplier = Decimal(str(selected_bullet['base_damage_multiplier']))
    base_armor_multiplier = Decimal(str(selected_bullet['base_armor_multiplier']))
    
    debug_print(f"子弹穿透等级: {penetration_level}, 伤害倍率: {base_damage_multiplier}, 护甲倍率: {base_armor_multiplier}")
    
    # 设置部位倍率
    body_part_multipliers = {
        '头部': Decimal(str(selected_weapon['head_multiplier'])),
        '胸部': Decimal(str(selected_weapon['chest_multiplier'])),
        '腹部': Decimal(str(selected_weapon['abdomen_multiplier'])),
        '大臂': Decimal(str(selected_weapon['upper_arm_multiplier'])),
        '小臂': Decimal(str(selected_weapon['lower_arm_multiplier'])),
        '大腿': Decimal(str(selected_weapon['thigh_multiplier'])),
        '小腿': Decimal(str(selected_weapon['calf_multiplier'])),
    }
    body_part_multipliers['下腹部'] = body_part_multipliers['腹部']
    
    debug_print("部位倍率:")
    for part, multiplier in body_part_multipliers.items():
        debug_print(f"  {part}: {multiplier}")
    
    # 打印选择的武器和子弹信息
    print("\n=== 武器信息 ===")
    print(f"武器: {selected_weapon['name']}")
    print(f"子弹: {selected_bullet['name']}")
    print(f"口径: {selected_weapon['raw_caliber']}")
    print(f"距离: {distance}米, 武器衰减倍率: {weapon_decay_multiplier}")
    print(f"穿透等级: {penetration_level}, 伤害倍率: {base_damage_multiplier}")
    print(f"护甲倍率: {base_armor_multiplier}")
    print(f"同级穿透倍率: {selected_bullet['same_level_penetration']}")
    print(f"越级穿透倍率: {selected_bullet['higher_level_penetration']}")
    print(f"头盔衰减倍率: {helmet_decay_multiplier}")
    print(f"护甲衰减倍率: {armor_decay_multiplier}")
    
    # 计算射击间隔
    if fire_rate:
        shot_interval = (Decimal('60000') / Decimal(str(fire_rate))).quantize(
            Decimal('0.01'), rounding=ROUND_HALF_UP
        )
    else:
        print("警告: 射速为0，使用默认值600")
        shot_interval = (Decimal('60000') / Decimal('600')).quantize(
            Decimal('0.01'), rounding=ROUND_HALF_UP
        )
    
    debug_print(f"射击间隔: {shot_interval} ms")
    
    # 初始化状态
    player_health = Decimal('100.0')
    current_helmet_durability = helmet_durability
    current_armor_durability = armor_durability
    valid_parts = ['头部', '胸部', '腹部', '大臂', '小臂', '大腿', '小腿', '下腹部', '未命中']
    
    hit_count = 0
    total_time = Decimal('0.0')
    total_damage = Decimal('0.0')
    total_armor_damage = Decimal('0.0')
    hit_statistics = {part: 0 for part in valid_parts}
    
    print("\n=== 开始模拟计算 ===")
    while True:
        # 输入命中部位
        while True:
            hit_part = input("\n输入命中部位 (头部/胸部/腹部/下腹部/大臂/小臂/大腿/小腿/未命中)：").strip()
            if hit_part.lower() == 'exit':
                return True
            if hit_part in valid_parts:
                hit_statistics[hit_part] += 1
                hit_count += 1
                break
            print("无效输入，请重新输入。")
        
        debug_print(f"命中部位: {hit_part}, 命中次数: {hit_count}")
        
        # 计算总耗时
        trigger_delay_dec = Decimal(str(trigger_delay))
        if fire_mode == 1:  # 全自动
            total_time = trigger_delay_dec + shot_interval * Decimal(str(hit_count - 1))
        else:  # 半自动
            total_time = (trigger_delay_dec * Decimal(str(hit_count))) + (shot_interval * Decimal(str(hit_count - 1)))
        total_time = total_time.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        debug_print(f"总耗时: {total_time} ms")
        
        # 处理未命中
        if hit_part == '未命中':
            print("\n=== 计算结果 ===")
            print("本次攻击未命中")
            print(f"累计耗时：{total_time} ms")
            continue
        
        # 伤害计算逻辑
        is_protected = False
        protector_level = 0
        protector_type = None
        current_protector_durability = Decimal('0.0')
        
        # 判断保护状态
        if hit_part == '头部':
            if helmet_level > 0 and current_helmet_durability > Decimal('0'):
                is_protected = True
                protector_level = helmet_level
                protector_type = 'helmet'
                current_protector_durability = current_helmet_durability
                debug_print("头部受头盔保护")
        else:
            if armor_level > 0 and current_armor_durability > Decimal('0') and hit_part in protected_areas:
                is_protected = True
                protector_level = armor_level
                protector_type = 'armor'
                current_protector_durability = current_armor_durability
                debug_print(f"{hit_part} 受护甲保护")
        
        # 计算穿透倍率（使用新的函数）
        penetration_multiplier = Decimal('0.0')
        if is_protected:
            penetration_multiplier = calculate_penetration_multiplier(
                penetration_level, protector_level, selected_bullet
            )
            debug_print(f"穿透倍率: {penetration_multiplier}")
        
        # 计算护甲伤害
        weapon_armor_damage_dec = Decimal(str(weapon_armor_damage))
        armor_damage_value = Decimal('0.0')
        final_damage = Decimal('0.0')
        protector_destroyed = False
        armor_damage_dealt = Decimal('0.0')
        
        # 特殊处理：.338 Lap Mag弹药始终完全穿透护甲
        is_338_lap_mag = selected_bullet['caliber'] == '338lapmag'
        
        if is_protected:
            # 根据防护装备类型选择正确的衰减倍率
            if protector_type == 'helmet':
                armor_damage_value = weapon_armor_damage_dec * base_armor_multiplier * helmet_decay_multiplier * weapon_decay_multiplier
            else:
                armor_damage_value = weapon_armor_damage_dec * base_armor_multiplier * armor_decay_multiplier * weapon_decay_multiplier
            
            debug_print(f"护甲伤害值: {armor_damage_value}")
            
            # 计算剩余耐久
            remaining_durability = current_protector_durability - armor_damage_value
            if remaining_durability <= Decimal('0'):
                protector_destroyed = True
                remaining_durability = Decimal('0.0')
            else:
                remaining_durability = remaining_durability.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
            
            debug_print(f"剩余耐久: {remaining_durability}, 护甲是否被击碎: {protector_destroyed}")
            
            # 累计护甲伤害
            armor_damage_dealt = current_protector_durability - remaining_durability
            total_armor_damage += armor_damage_dealt
            
            # 计算伤害
            part_multiplier = body_part_multipliers[hit_part]
            denominator = weapon_armor_damage_dec * base_armor_multiplier * weapon_decay_multiplier
            
            if denominator == Decimal('0'):
                ratio = Decimal('0.0')
            else:
                # 根据防护类型使用正确的衰减倍率
                if protector_type == 'helmet':
                    denominator *= helmet_decay_multiplier
                else:
                    denominator *= armor_decay_multiplier
                ratio = current_protector_durability / denominator
            
            debug_print(f"部位倍率: {part_multiplier}, 分母: {denominator}, 比率: {ratio}")
            
            weapon_damage_dec = Decimal(str(weapon_damage))
            
            # 特殊处理：.338弹药完全穿透护甲
            if is_338_lap_mag:
                # .338弹药完全穿透护甲，直接造成全额伤害
                final_damage = weapon_damage_dec * base_damage_multiplier * part_multiplier * weapon_decay_multiplier
                print("\n[.338 Lap Mag特殊效果] 完全穿透护甲！")
                debug_print(".338 Lap Mag特殊效果: 完全穿透护甲")
            else:
                # 正常护甲穿透计算
                if current_protector_durability >= armor_damage_value:
                    final_damage = weapon_damage_dec * base_damage_multiplier * part_multiplier * penetration_multiplier * weapon_decay_multiplier
                    debug_print("护甲未被击穿，使用穿透倍率计算伤害")
                else:
                    part1 = ratio * weapon_damage_dec * base_damage_multiplier * part_multiplier * penetration_multiplier * weapon_decay_multiplier
                    part2 = (Decimal('1') - ratio) * weapon_damage_dec * base_damage_multiplier * part_multiplier * weapon_decay_multiplier
                    final_damage = part1 + part2
                    debug_print(f"护甲部分击穿，伤害分两部分计算: {part1} + {part2}")
            
            # 四舍五入伤害值
            final_damage = final_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            # 更新耐久
            if protector_type == 'helmet':
                current_helmet_durability = remaining_durability
            else:
                current_armor_durability = remaining_durability
        else:
            # 未受保护
            part_multiplier = body_part_multipliers[hit_part]
            weapon_damage_dec = Decimal(str(weapon_damage))
            final_damage = weapon_damage_dec * base_damage_multiplier * part_multiplier * weapon_decay_multiplier
            final_damage = final_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            debug_print("无保护，直接计算伤害")
        
        debug_print(f"最终伤害: {final_damage}")
        
        total_damage += final_damage
        player_health -= final_damage
        
        # 输出结果
        print("\n=== 计算结果 ===")
        if is_protected and not is_338_lap_mag:  # .338弹药不显示护甲阻挡信息
            if protector_type == 'helmet':
                print("头盔被击碎！" if protector_destroyed else "头盔未被击碎。")
                print(f"头盔损失耐久：{armor_damage_dealt.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}")
            else:
                print("护甲被击碎！" if protector_destroyed else "护甲未被击碎。")
                print(f"护甲损失耐久：{armor_damage_dealt.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}")
        elif hit_part == '头部' and helmet_level > 0 and not is_338_lap_mag:
            print("（未受头盔保护）")
        
        print(f"受到伤害：{final_damage}")
        print(f"剩余生命值：{player_health.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}")
        print(f"剩余头盔耐久：{current_helmet_durability}")
        print(f"剩余护甲耐久：{current_armor_durability}")
        print(f"累计耗时：{total_time} ms")
        
        # 死亡处理
        if player_health <= Decimal('0'):
            print("\n=== 最终统计 ===")
            print(f"射击模式：{'全自动' if fire_mode == 1 else '半自动'}")
            print(f"总造成伤害：{total_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}")
            print(f"总护甲伤害：{total_armor_damage.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}")
            
            print("\n命中统计：")
            total_shots = sum(hit_statistics.values())
            valid_hits = total_shots - hit_statistics['未命中']
            display_order = ['未命中', '头部', '胸部', '腹部', '下腹部', '大臂', '小臂', '大腿', '小腿']
            for part in display_order:
                count = hit_statistics.get(part, 0)
                if count > 0:
                    print(f"{part.ljust(5)}：{count}次")
            
            print(f"\n有效命中次数：{valid_hits}次")
            print(f"总攻击次数：{total_shots}次")
            print(f"击杀耗时：{total_time} ms")
            
            # 使用msvcrt等待按键
            print("\n按任意键结束本次模拟计算...")
            msvcrt.getch()
            return True

def main():
    global DEBUG_MODE
    
    print("三角洲行动夺金伤害计算模拟程序 V0.2.9")  # 版本号更新
    print("按 ESC 键可随时退出程序")
    
    # 检查是否启用调试模式
    if len(sys.argv) > 1 and (sys.argv[1] == "--debug" or sys.argv[1] == "-d"):
        DEBUG_MODE = True
        print("调试模式已启用")
    
    while True:
        result = run_simulation()
        if not result:
            break
            
        # 使用msvcrt检测按键
        if not wait_for_key():
            print("\n感谢使用，再见！")
            break
            
        # 清屏并开始新一轮计算
        os.system('cls' if os.name == 'nt' else 'clear')
        print("开始新一轮计算...")

if __name__ == "__main__":
    main()        
        for row in range(3, ws.max_row + 1):
            # 检查是否是新的武器类别
            category_cell = ws.cell(row=row, column=1).value
            if category_cell and str(category_cell).strip():
                current_category = str(category_cell).strip()
                continue
                
            weapon_name = ws.cell(row=row, column=2).value
            if not weapon_name or not str(weapon_name).strip():
                continue
                
            # 标准化口径
            caliber = ws.cell(row=row, column=3).value
            std_caliber = standardize_caliber(caliber) if caliber else ""
            
            # 创建武器数据字典
            weapon_data = {
                'category': current_category,
                'name': str(weapon_name).strip(),
                'caliber': std_caliber,
                'raw_caliber': caliber,  # 保留原始口径用于显示
                'fire_mode': ws.cell(row=row, column=5).value or 0,  # E列 (射击模式-程序用)
                'trigger_delay': ws.cell(row=row, column=6).value or 0,  # F列 (扳机延迟)
                'fire_rate': ws.cell(row=row, column=7).value or 0,  # G列 (射速)
                'muzzle_velocity': ws.cell(row=row, column=8).value or 0,  # H列 (枪口初速)
                'base_damage': ws.cell(row=row, column=9).value or 0,  # I列 (基础伤害)
                'armor_damage': ws.cell(row=row, column=11).value or 0,  # K列 (护甲伤害)
                # 部位倍率
                'head_multiplier': ws.cell(row=row, column=13).value or 1.0,  # M列
                'chest_multiplier': ws.cell(row=row, column=14).value or 1.0,  # N列
                'abdomen_multiplier': ws.cell(row=row, column=15).value or 1.0,  # O列
                'upper_arm_multiplier': ws.cell(row=row, column=16).value or 1.0,  # P列
                'lower_arm_multiplier': ws.cell(row=row, column=17).value or 1.0,  # Q列
                'thigh_multiplier': ws.cell(row=row, column=18).value or 1.0,  # R列
                'calf_multiplier': ws.cell(row=row, column=19).value or 1.0,  # S列
                'decay_distances': [],
                'decay_factors': []
            }
            
            # 读取衰减数据 (最多4组)
            decay_cols = [(20, 21), (22, 23), (24, 25), (26, 27)]  # T/U, V/W, X/Y, Z/AA列
            
            for dist_col, factor_col in decay_cols:
                dist = ws.cell(row=row, column=dist_col).value
                factor = ws.cell(row=row, column=factor_col).value
                
                # 处理特殊值
                if dist in (None, "", "/", "//", "N/A"):
                    continue
                    
                try:
                    # 尝试转换为数值
                    dist_val = float(dist)
                    factor_val = float(factor)
                    weapon_data['decay_distances'].append(dist_val)
                    weapon_data['decay_factors'].append(factor_val)
                except (ValueError, TypeError):
                    continue
            
            weapons.append(weapon_data)
        
        debug_print(f"成功加载 {len(weapons)} 种武器")
        return weapons
    
    except Exception as e:
        print(f"\n错误: 无法加载武器数据 - {e}")
        return []

def load_bullets_data():
    """加载子弹数据"""
    try:
        wb = openpyxl.load_workbook('S6子弹数据.xlsx', data_only=True)
        ws = wb['子弹数据']
        
        bullets = []
        current_caliber = ""
        
        for row in range(3, ws.max_row + 1):
            # 检查是否是新的子弹口径
            caliber_cell = ws.cell(row=row, column=1).value
            if caliber_cell and str(caliber_cell).strip():
                current_caliber = standardize_caliber(caliber_cell)
                
            bullet_name = ws.cell(row=row, column=2).value
            if not bullet_name or not str(bullet_name).strip():
                continue
                
            # 处理可能的空值
            pellet_count = ws.cell(row=row, column=3).value or 1
            penetration_level = ws.cell(row=row, column=4).value or 0
            base_damage_multiplier = ws.cell(row=row, column=5).value or 1.0
            base_armor_multiplier = ws.cell(row=row, column=6).value or 1.0
            
            bullet_data = {
                'caliber': current_caliber,
                'name': str(bullet_name).strip(),
                'pellet_count': pellet_count,
                'penetration_level': penetration_level,
                'base_damage_multiplier': base_damage_multiplier,
                'base_armor_multiplier': base_armor_multiplier,
                'armor_decay_factors': []
            }
            
            # 读取护甲衰减倍率 (1-6级)
            for col in range(7, 13):  # G到L列
                factor = ws.cell(row=row, column=col).value
                if factor is None:
                    factor = 0.0
                try:
                    bullet_data['armor_decay_factors'].append(float(factor))
                except (ValueError, TypeError):
                    bullet_data['armor_decay_factors'].append(0.0)
            
            bullets.append(bullet_data)
        
        debug_print(f"成功加载 {len(bullets)} 种子弹")
        return bullets
    
    except Exception as e:
        print(f"\n错误: 无法加载子弹数据 - {e}")
        return []

def calculate_weapon_decay(distance, weapon):
    """计算武器衰减倍率"""
    if not weapon['decay_distances']:
        debug_print(f"武器 {weapon['name']} 无衰减数据，使用默认倍率 1.0")
        return Decimal('1.0')
    
    # 如果距离小于第一个衰减距离，无衰减
    if distance <= weapon['decay_distances'][0]:
        debug_print(f"距离 {distance} 小于第一个衰减距离 {weapon['decay_distances'][0]}，无衰减")
        return Decimal('1.0')
    
    # 检查后续衰减距离
    for i in range(1, len(weapon['decay_distances'])):
        if distance <= weapon['decay_distances'][i]:
            decay_factor = Decimal(str(weapon['decay_factors'][i-1]))
            debug_print(f"距离 {distance} 在衰减区间 {weapon['decay_distances'][i-1]}-{weapon['decay_distances'][i]}，衰减倍率: {decay_factor}")
            return decay_factor
    
    # 如果超过所有衰减距离，使用最后一个衰减倍率
    decay_factor = Decimal(str(weapon['decay_factors'][-1]))
    debug_print(f"距离 {distance} 超过所有衰减距离，使用最后衰减倍率: {decay_factor}")
    return decay_factor

# 读取护甲与头盔数据
def load_armor_data():
    """加载护甲和头盔数据"""
    try:
        wb = openpyxl.load_workbook('S6护甲数据.xlsx', data_only=True)
        ws = wb.active
        
        armors = []
        helmets = []
        current_section = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            
            # 检测章节开始
            if cell_value == "护甲":
                current_section = "armor"
                continue
            elif cell_value == "头盔":
                current_section = "helmet"
                continue
                
            if not cell_value or not str(cell_value).strip():
                continue
                
            # 读取数据
            name = str(cell_value).strip()
            level = ws.cell(row=row, column=2).value
            armor_type = ws.cell(row=row, column=3).value
            max_durability = ws.cell(row=row, column=7).value  # G列: 初始上限
            
            if level is None or max_durability is None:
                continue
                
            try:
                level = int(level)
                max_durability = Decimal(str(max_durability))
            except (ValueError, TypeError):
                continue
                
            item_data = {
                'name': name,
                'level': level,
                'max_durability': max_durability
            }
            
            if current_section == "armor":
                # 护甲类型映射: 半甲->1, 全甲->2, 重甲->3
                if armor_type == '半甲':
                    item_data['armor_type'] = 1
                elif armor_type == '全甲':
                    item_data['armor_type'] = 2
                elif armor_type == '重甲':
                    item_data['armor_type'] = 3
                else:
                    continue  # 跳过无效类型
                armors.append(item_data)
                
            elif current_section == "helmet":
                helmets.append(item_data)
                
        debug_print(f"成功加载 {len(armors)} 种护甲和 {len(helmets)} 种头盔")
        return armors, helmets
    
    except Exception as e:
        print(f"\n错误: 无法加载护甲数据 - {e}")
        return [], []

def select_protection(items, item_type):
    """选择防护装备并输入耐久"""
    print(f"\n=== 选择{item_type} ===")
    print("0. 无")
    
    # 按等级分组
    levels = {}
    for item in items:
        level = item['level']
        if level not in levels:
            levels[level] = []
        levels[level].append(item)
    
    # 显示等级选项
    sorted_levels = sorted(levels.keys())
    for i, level in enumerate(sorted_levels, 1):
        print(f"{i}. {level}级{item_type}")
    
    # 获取用户选择，包括0（无）选项
    max_choice = len(sorted_levels)
    level_choice = get_int_input(f"请选择{item_type}等级 (0-{max_choice}): ", 0, max_choice)
    
    # 处理"无"选项
    if level_choice == 0:
        debug_print(f"用户选择无{item_type}")
        return None, Decimal('0.0')
    
    selected_level = sorted_levels[level_choice - 1]
    
    # 显示该等级下的装备
    level_items = levels[selected_level]
    print(f"\n{selected_level}级{item_type}列表:")
    for i, item in enumerate(level_items, 1):
        print(f"{i}. {item['name']} (最大耐久: {item['max_durability']})")
    
    item_choice = get_int_input(f"请选择{item_type}: ", 1, len(level_items))
    selected_item = level_items[item_choice - 1]
    
    # 输入当前耐久
    max_durability = float(selected_item['max_durability'])
    durability = get_decimal_input(
        f"请输入当前{item_type}耐久 (0.0-{max_durability}): ",
        0.0, max_durability, 1
    )
    
    debug_print(f"选择{item_type}: {selected_item['name']}, 等级: {selected_item['level']}, 耐久: {durability}")
    return selected_item, durability

def run_simulation():
    """运行一次完整的伤害模拟"""
    # 初始化参数
    print("=== 通用武器伤害模拟器 ===")
    
    # 加载护甲数据
    print("正在加载护甲数据...")
    armors, helmets = load_armor_data()
    
    # 选择头盔
    helmet_level = 0
    helmet_durability = Decimal('0.0')
    
    if helmets:
        selected_helmet, helmet_durability = select_protection(helmets, "头盔")
        if selected_helmet:
            helmet_level = selected_helmet['level']
            print(f"已选择头盔: {selected_helmet['name']} (等级{helmet_level})")
        else:
            print("已选择：无头盔")
    else:
        print("未找到头盔数据，将使用无头盔设置")
    
    # 选择护甲
    armor_level = 0
    armor_durability = Decimal('0.0')
    armor_type_value = 0
    
    if armors:
        selected_armor, armor_durability = select_protection(armors, "护甲")
        if selected_armor:
            armor_level = selected_armor['level']
            armor_type_value = selected_armor['armor_type']
            print(f"已选择护甲: {selected_armor['name']} (等级{armor_level}, 类型{armor_type_value})")
        else:
            print("已选择：无护甲")
    else:
        print("未找到护甲数据，将使用无护甲设置")
    
    # 设置保护的身体部位
    protected_areas = {
        1: ['胸部', '腹部'],
        2: ['胸部', '腹部', '下腹部'],
        3: ['胸部', '腹部', '下腹部', '大臂']
    }.get(armor_type_value, [])  # 如果没有护甲，返回空列表
    
    debug_print(f"护甲类型: {armor_type_value}, 保护部位: {protected_areas}")
    
    # 加载武器和子弹数据
    print("正在加载武器数据...")
    weapons = load_weapons_data()
    if not weapons:
        print("无法加载武器数据，程序退出")
        return False
    
    print("正在加载子弹数据...")
    bullets = load_bullets_data()
    if not bullets:
        print("无法加载子弹数据，程序退出")
        return False
    
    # 选择武器
    print("\n=== 选择武器 ===")
    print("（点射武器与多弹丸武器不适用本程序）")
    
    # 按类别分组武器
    categories = {}
    for weapon in weapons:
        if weapon['category'] not in categories:
            categories[weapon['category']] = []
        categories[weapon['category']].append(weapon)
    
    # 显示武器类别
    print("\n请选择武器类别：")
    category_list = list(categories.keys())
    for i, category in enumerate(category_list, 1):
        print(f"{i}. {category}")
    
    category_choice = get_int_input("输入类别编号：", 1, len(category_list))
    selected_category = category_list[category_choice - 1]
    
    # 显示选定类别中的武器
    print(f"\n{selected_category}武器列表：")
    category_weapons = categories[selected_category]
    for i, weapon in enumerate(category_weapons, 1):
        print(f"{i}. {weapon['name']} ({weapon['raw_caliber']})")
    
    weapon_choice = get_int_input("输入武器编号：", 1, len(category_weapons))
    selected_weapon = category_weapons[weapon_choice - 1]
    weapon_caliber = selected_weapon['caliber']
    
    debug_print(f"选择武器: {selected_weapon['name']}, 口径: {weapon_caliber}")
    
    # 霰弹枪警告
    if selected_weapon['category'] == "霰弹枪":
        print("警告：霰弹枪伤害计算可能不准确（多弹丸特性）")
    
    # 根据武器口径过滤子弹
    caliber_bullets = [b for b in bullets if b['caliber'] == weapon_caliber]
    
    if not caliber_bullets:
        print(f"\n警告：没有找到匹配 {weapon_caliber} 口径的子弹！")
        print(f"武器口径: {selected_weapon['raw_caliber']}")
        print("可用的子弹口径:")
        unique_calibers = set(b['caliber'] for b in bullets)
        for cal in unique_calibers:
            print(f" - {cal}")
        print("程序将退出。")
        return False
    
    # 选择子弹
    print(f"\n=== 选择 {selected_weapon['raw_caliber']} 子弹 ===")
    for i, bullet in enumerate(caliber_bullets, 1):
        pen_level = bullet['penetration_level']
        dmg_mult = bullet['base_damage_multiplier']
        print(f"{i}. {bullet['name']} (穿透等级:{pen_level}, 伤害倍率:{dmg_mult})")
    
    bullet_choice = get_int_input("输入子弹编号：", 1, len(caliber_bullets))
    selected_bullet = caliber_bullets[bullet_choice - 1]
    
    debug_print(f"选择子弹: {selected_bullet['name']}, 穿透等级: {selected_bullet['penetration_level']}, 伤害倍率: {selected_bullet['base_damage_multiplier']}")
    
    # 输入目标距离
    distance = get_decimal_input("\n请输入目标距离（0-400米）：", 0.0, 400.0, 1)
    
    # 计算武器衰减倍率
    weapon_decay_multiplier = calculate_weapon_decay(float(distance), selected_weapon)
    
    # 初始化头盔和护甲衰减倍率
    helmet_decay_multiplier = Decimal('0.0')
    armor_decay_multiplier = Decimal('0.0')
    
    # 计算头盔衰减倍率 (根据头盔等级)
    if helmet_level > 0 and helmet_level <= 6:
        helmet_decay_multiplier = Decimal(str(
            selected_bullet['armor_decay_factors'][helmet_level - 1]
        ))
        debug_print(f"头盔等级 {helmet_level}, 头盔衰减倍率: {helmet_decay_multiplier}")
    
    # 计算护甲衰减倍率 (根据护甲等级)
    if armor_level > 0 and armor_level <= 6:
        armor_decay_multiplier = Decimal(str(
            selected_bullet['armor_decay_factors'][armor_level - 1]
        ))
        debug_print(f"护甲等级 {armor_level}, 护甲衰减倍率: {armor_decay_multiplier}")
    
    # 设置武器参数
    weapon_damage = selected_weapon['base_damage']
    weapon_armor_damage = selected_weapon['armor_damage']
    fire_rate = selected_weapon['fire_rate']
    fire_mode = selected_weapon['fire_mode']
    trigger_delay = selected_weapon['trigger_delay']
    
    debug_print(f"武器基础伤害: {weapon_damage}, 武器护甲伤害: {weapon_armor_damage}")
    debug_print(f"射速: {fire_rate}, 射击模式: {fire_mode}, 扳机延迟: {trigger_delay}")
    
    # 设置子弹参数
    penetration_level = selected_bullet['penetration_level']
    base_damage_multiplier = Decimal(str(selected_bullet['base_damage_multiplier']))
    base_armor_multiplier = Decimal(str(selected_bullet['base_armor_multiplier']))
    
    debug_print(f"子弹穿透等级: {penetration_level}, 伤害倍率: {base_damage_multiplier}, 护甲倍率: {base_armor_multiplier}")
    
    # 设置部位倍率
    body_part_multipliers = {
        '头部': Decimal(str(selected_weapon['head_multiplier'])),
        '胸部': Decimal(str(selected_weapon['chest_multiplier'])),
        '腹部': Decimal(str(selected_weapon['abdomen_multiplier'])),
        '大臂': Decimal(str(selected_weapon['upper_arm_multiplier'])),
        '小臂': Decimal(str(selected_weapon['lower_arm_multiplier'])),
        '大腿': Decimal(str(selected_weapon['thigh_multiplier'])),
        '小腿': Decimal(str(selected_weapon['calf_multiplier'])),
    }
    body_part_multipliers['下腹部'] = body_part_multipliers['腹部']
    
    debug_print("部位倍率:")
    for part, multiplier in body_part_multipliers.items():
        debug_print(f"  {part}: {multiplier}")
    
    # 打印选择的武器和子弹信息
    print("\n=== 武器信息 ===")
    print(f"武器: {selected_weapon['name']}")
    print(f"子弹: {selected_bullet['name']}")
    print(f"口径: {selected_weapon['raw_caliber']}")
    print(f"距离: {distance}米, 武器衰减倍率: {weapon_decay_multiplier}")
    print(f"穿透等级: {penetration_level}, 伤害倍率: {base_damage_multiplier}")
    print(f"护甲倍率: {base_armor_multiplier}")
    print(f"头盔衰减倍率: {helmet_decay_multiplier}")
    print(f"护甲衰减倍率: {armor_decay_multiplier}")
    
    # 计算射击间隔
    if fire_rate:
        shot_interval = (Decimal('60000') / Decimal(str(fire_rate))).quantize(
            Decimal('0.01'), rounding=ROUND_HALF_UP
        )
    else:
        print("警告: 射速为0，使用默认值600")
        shot_interval = (Decimal('60000') / Decimal('600')).quantize(
            Decimal('0.01'), rounding=ROUND_HALF_UP
        )
    
    debug_print(f"射击间隔: {shot_interval} ms")
    
    # 初始化状态
    player_health = Decimal('100.0')
    current_helmet_durability = helmet_durability
    current_armor_durability = armor_durability
    valid_parts = ['头部', '胸部', '腹部', '大臂', '小臂', '大腿', '小腿', '下腹部', '未命中']
    
    hit_count = 0
    total_time = Decimal('0.0')
    total_damage = Decimal('0.0')
    total_armor_damage = Decimal('0.0')
    hit_statistics = {part: 0 for part in valid_parts}
    
    print("\n=== 开始模拟计算 ===")
    while True:
        # 输入命中部位
        while True:
            hit_part = input("\n输入命中部位 (头部/胸部/腹部/下腹部/大臂/小臂/大腿/小腿/未命中)：").strip()
            if hit_part.lower() == 'exit':
                return True
            if hit_part in valid_parts:
                hit_statistics[hit_part] += 1
                hit_count += 1
                break
            print("无效输入，请重新输入。")
        
        debug_print(f"命中部位: {hit_part}, 命中次数: {hit_count}")
        
        # 计算总耗时
        trigger_delay_dec = Decimal(str(trigger_delay))
        if fire_mode == 1:  # 全自动
            total_time = trigger_delay_dec + shot_interval * Decimal(str(hit_count - 1))
        else:  # 半自动
            total_time = (trigger_delay_dec * Decimal(str(hit_count))) + (shot_interval * Decimal(str(hit_count - 1)))
        total_time = total_time.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        debug_print(f"总耗时: {total_time} ms")
        
        # 处理未命中
        if hit_part == '未命中':
            print("\n=== 计算结果 ===")
            print("本次攻击未命中")
            print(f"累计耗时：{total_time} ms")
            continue
        
        # 伤害计算逻辑
        is_protected = False
        protector_level = 0
        protector_type = None
        current_protector_durability = Decimal('0.0')
        
        # 判断保护状态
        if hit_part == '头部':
            if helmet_level > 0 and current_helmet_durability > Decimal('0'):
                is_protected = True
                protector_level = helmet_level
                protector_type = 'helmet'
                current_protector_durability = current_helmet_durability
                debug_print("头部受头盔保护")
        else:
            if armor_level > 0 and current_armor_durability > Decimal('0') and hit_part in protected_areas:
                is_protected = True
                protector_level = armor_level
                protector_type = 'armor'
                current_protector_durability = current_armor_durability
                debug_print(f"{hit_part} 受护甲保护")
        
        # 计算穿透倍率
        penetration_multiplier = Decimal('0.0')
        if is_protected:
            diff = penetration_level - protector_level
            if diff < 0:
                penetration_multiplier = Decimal('0.0')
            elif diff == 0:
                penetration_multiplier = Decimal('0.5')
            elif diff == 1:
                penetration_multiplier = Decimal('0.75')
            else:
                penetration_multiplier = Decimal('1.0')
            debug_print(f"穿透等级差: {diff}, 穿透倍率: {penetration_multiplier}")
        
        # 计算护甲伤害
        weapon_armor_damage_dec = Decimal(str(weapon_armor_damage))
        armor_damage_value = Decimal('0.0')
        final_damage = Decimal('0.0')
        protector_destroyed = False
        armor_damage_dealt = Decimal('0.0')
        
        # 特殊处理：.338 Lap Mag弹药始终完全穿透护甲
        is_338_lap_mag = selected_bullet['caliber'] == '338lapmag'
        
        if is_protected:
            # 根据防护装备类型选择正确的衰减倍率
            if protector_type == 'helmet':
                armor_damage_value = weapon_armor_damage_dec * base_armor_multiplier * helmet_decay_multiplier * weapon_decay_multiplier
            else:
                armor_damage_value = weapon_armor_damage_dec * base_armor_multiplier * armor_decay_multiplier * weapon_decay_multiplier
            
            debug_print(f"护甲伤害值: {armor_damage_value}")
            
            # 计算剩余耐久
            remaining_durability = current_protector_durability - armor_damage_value
            if remaining_durability <= Decimal('0'):
                protector_destroyed = True
                remaining_durability = Decimal('0.0')
            else:
                remaining_durability = remaining_durability.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
            
            debug_print(f"剩余耐久: {remaining_durability}, 护甲是否被击碎: {protector_destroyed}")
            
            # 累计护甲伤害
            armor_damage_dealt = current_protector_durability - remaining_durability
            total_armor_damage += armor_damage_dealt
            
            # 计算伤害
            part_multiplier = body_part_multipliers[hit_part]
            denominator = weapon_armor_damage_dec * base_armor_multiplier * weapon_decay_multiplier
            
            if denominator == Decimal('0'):
                ratio = Decimal('0.0')
            else:
                # 根据防护类型使用正确的衰减倍率
                if protector_type == 'helmet':
                    denominator *= helmet_decay_multiplier
                else:
                    denominator *= armor_decay_multiplier
                ratio = current_protector_durability / denominator
            
            debug_print(f"部位倍率: {part_multiplier}, 分母: {denominator}, 比率: {ratio}")
            
            weapon_damage_dec = Decimal(str(weapon_damage))
            
            # 特殊处理：.338弹药完全穿透护甲
            if is_338_lap_mag:
                # .338弹药完全穿透护甲，直接造成全额伤害
                final_damage = weapon_damage_dec * base_damage_multiplier * part_multiplier * weapon_decay_multiplier
                print("\n[.338 Lap Mag特殊效果] 完全穿透护甲！")
                debug_print(".338 Lap Mag特殊效果: 完全穿透护甲")
            else:
                # 正常护甲穿透计算
                if current_protector_durability >= armor_damage_value:
                    final_damage = weapon_damage_dec * base_damage_multiplier * part_multiplier * penetration_multiplier * weapon_decay_multiplier
                    debug_print("护甲未被击穿，使用穿透倍率计算伤害")
                else:
                    part1 = ratio * weapon_damage_dec * base_damage_multiplier * part_multiplier * penetration_multiplier * weapon_decay_multiplier
                    part2 = (Decimal('1') - ratio) * weapon_damage_dec * base_damage_multiplier * part_multiplier * weapon_decay_multiplier
                    final_damage = part1 + part2
                    debug_print(f"护甲部分击穿，伤害分两部分计算: {part1} + {part2}")
            
            # 四舍五入伤害值
            final_damage = final_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            # 更新耐久
            if protector_type == 'helmet':
                current_helmet_durability = remaining_durability
            else:
                current_armor_durability = remaining_durability
        else:
            # 未受保护
            part_multiplier = body_part_multipliers[hit_part]
            weapon_damage_dec = Decimal(str(weapon_damage))
            final_damage = weapon_damage_dec * base_damage_multiplier * part_multiplier * weapon_decay_multiplier
            final_damage = final_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            debug_print("无保护，直接计算伤害")
        
        debug_print(f"最终伤害: {final_damage}")
        
        total_damage += final_damage
        player_health -= final_damage
        
        # 输出结果
        print("\n=== 计算结果 ===")
        if is_protected and not is_338_lap_mag:  # .338弹药不显示护甲阻挡信息
            if protector_type == 'helmet':
                print("头盔被击碎！" if protector_destroyed else "头盔未被击碎。")
                print(f"头盔损失耐久：{armor_damage_dealt.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}")
            else:
                print("护甲被击碎！" if protector_destroyed else "护甲未被击碎。")
                print(f"护甲损失耐久：{armor_damage_dealt.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}")
        elif hit_part == '头部' and helmet_level > 0 and not is_338_lap_mag:
            print("（未受头盔保护）")
        
        print(f"受到伤害：{final_damage}")
        print(f"剩余生命值：{player_health.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}")
        print(f"剩余头盔耐久：{current_helmet_durability}")
        print(f"剩余护甲耐久：{current_armor_durability}")
        print(f"累计耗时：{total_time} ms")
        
        # 死亡处理
        if player_health <= Decimal('0'):
            print("\n=== 最终统计 ===")
            print(f"射击模式：{'全自动' if fire_mode == 1 else '半自动'}")
            print(f"总造成伤害：{total_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}")
            print(f"总护甲伤害：{total_armor_damage.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}")
            
            print("\n命中统计：")
            total_shots = sum(hit_statistics.values())
            valid_hits = total_shots - hit_statistics['未命中']
            display_order = ['未命中', '头部', '胸部', '腹部', '下腹部', '大臂', '小臂', '大腿', '小腿']
            for part in display_order:
                count = hit_statistics.get(part, 0)
                if count > 0:
                    print(f"{part.ljust(5)}：{count}次")
            
            print(f"\n有效命中次数：{valid_hits}次")
            print(f"总攻击次数：{total_shots}次")
            print(f"击杀耗时：{total_time} ms")
            
            # 使用msvcrt等待按键
            print("\n按任意键结束本次模拟计算...")
            msvcrt.getch()
            return True

def main():
    global DEBUG_MODE
    
    print("三角洲行动夺金伤害计算模拟程序 V0.2.8")
    print("按 ESC 键可随时退出程序")
    
    # 检查是否启用调试模式
    if len(sys.argv) > 1 and (sys.argv[1] == "--debug" or sys.argv[1] == "-d"):
        DEBUG_MODE = True
        print("调试模式已启用")
    
    while True:
        result = run_simulation()
        if not result:
            break
            
        # 使用msvcrt检测按键
        if not wait_for_key():
            print("\n感谢使用，再见！")
            break
            
        # 清屏并开始新一轮计算
        os.system('cls' if os.name == 'nt' else 'clear')
        print("开始新一轮计算...")

if __name__ == "__main__":

    main()
