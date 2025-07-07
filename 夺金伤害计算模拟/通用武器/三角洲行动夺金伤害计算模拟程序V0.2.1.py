print("本程序由B站繁星攻略组制作")

import sys
from decimal import Decimal, ROUND_HALF_UP
import openpyxl

def get_decimal_input(prompt, min_val, max_val, decimal_places):
    """获取精确的小数输入"""
    while True:
        try:
            value = input(prompt)
            decimal_value = Decimal(value)
            quantized = decimal_value.quantize(
                Decimal('1.' + '0' * decimal_places), 
                rounding=ROUND_HALF_UP
            )
            float_val = float(quantized)
            if min_val <= float_val <= max_val:
                return quantized
            else:
                print(f"输入错误，范围{min_val}到{max_val}，请重新输入。")
        except Exception:
            print("输入错误，请输入有效的数字。")

def get_int_input(prompt, min_val, max_val):
    """获取整数输入"""
    while True:
        try:
            value = int(input(prompt))
            if min_val <= value <= max_val:
                return value
            else:
                print(f"输入错误，范围{min_val}到{max_val}，请重新输入。")
        except ValueError:
            print("输入错误，请输入整数。")

def load_weapons_data():
    """加载武器数据"""
    wb = openpyxl.load_workbook('S5夺金武器.xlsx')
    ws = wb['夺金模式']
    
    weapons = []
    current_category = ""
    
    for row in range(3, ws.max_row + 1):
        # 检查是否是新的武器类别
        category_cell = ws.cell(row=row, column=1).value
        if category_cell:
            current_category = category_cell
            continue
            
        weapon_name = ws.cell(row=row, column=2).value
        if not weapon_name:
            continue
            
        weapon_data = {
            'category': current_category,
            'name': weapon_name,
            'caliber': ws.cell(row=row, column=3).value,
            'fire_mode': ws.cell(row=row, column=4).value,  # 射击模式(程序用)
            'trigger_delay': ws.cell(row=row, column=5).value,
            'fire_rate': ws.cell(row=row, column=6).value,
            'muzzle_velocity': ws.cell(row=row, column=7).value,
            'base_damage': ws.cell(row=row, column=8).value,
            'armor_damage': ws.cell(row=row, column=10).value,
            'head_multiplier': ws.cell(row=row, column=12).value,
            'chest_multiplier': ws.cell(row=row, column=13).value,
            'abdomen_multiplier': ws.cell(row=row, column=14).value,
            'upper_arm_multiplier': ws.cell(row=row, column=15).value,
            'lower_arm_multiplier': ws.cell(row=row, column=16).value,
            'thigh_multiplier': ws.cell(row=row, column=17).value,
            'calf_multiplier': ws.cell(row=row, column=18).value,
            'decay_distances': [],
            'decay_factors': []
        }
        
        # 读取衰减数据 (最多4组)
        decay_cols = [(19, 20), (21, 22), (23, 24), (25, 26)]
        for dist_col, factor_col in decay_cols:
            dist = ws.cell(row=row, column=dist_col).value
            factor = ws.cell(row=row, column=factor_col).value
            
            if dist and factor and dist != '/' and factor != '/':
                try:
                    weapon_data['decay_distances'].append(float(dist))
                    weapon_data['decay_factors'].append(float(factor))
                except (ValueError, TypeError):
                    continue
        
        weapons.append(weapon_data)
    
    return weapons

def load_bullets_data():
    """加载子弹数据"""
    wb = openpyxl.load_workbook('S5子弹数据.xlsx')
    ws = wb['子弹数据']
    
    bullets = []
    current_caliber = ""
    
    for row in range(3, ws.max_row + 1):
        # 检查是否是新的子弹口径
        caliber_cell = ws.cell(row=row, column=1).value
        if caliber_cell:
            current_caliber = caliber_cell
            continue
            
        bullet_name = ws.cell(row=row, column=2).value
        if not bullet_name:
            continue
            
        bullet_data = {
            'caliber': current_caliber,
            'name': bullet_name,
            'pellet_count': ws.cell(row=row, column=3).value,
            'penetration_level': ws.cell(row=row, column=4).value,
            'base_damage_multiplier': ws.cell(row=row, column=5).value,
            'base_armor_multiplier': ws.cell(row=row, column=6).value,
            'armor_decay_factors': []
        }
        
        # 读取护甲衰减倍率 (1-6级)
        for col in range(7, 13):
            factor = ws.cell(row=row, column=col).value
            if factor is None:
                factor = 0.0
            bullet_data['armor_decay_factors'].append(float(factor))
        
        bullets.append(bullet_data)
    
    return bullets

def calculate_weapon_decay(distance, weapon):
    """计算武器衰减倍率"""
    if not weapon['decay_distances']:
        return Decimal('1.0')
    
    # 如果距离小于第一个衰减距离，无衰减
    if distance <= weapon['decay_distances'][0]:
        return Decimal('1.0')
    
    # 检查后续衰减距离
    for i in range(len(weapon['decay_distances'])):
        if distance <= weapon['decay_distances'][i]:
            return Decimal(str(weapon['decay_factors'][i-1]))
    
    # 如果超过所有衰减距离，使用最后一个衰减倍率
    return Decimal(str(weapon['decay_factors'][-1]))

def main():
    # 初始化参数
    print("=== 初始化参数 ===")
    helmet_durability = get_decimal_input("请输入头盔耐久（0.0-60.0）：", 0.0, 60.0, 1)
    armor_durability = get_decimal_input("请输入护甲耐久（0.0-150.0）：", 0.0, 150.0, 1)
    helmet_level = get_int_input("请输入头盔防护等级（0-6）：", 0, 6)
    armor_level = get_int_input("请输入护甲防护等级（0-6）：", 0, 6)
    armor_type = get_int_input("\n请输入护甲类型（1-半甲，2-全甲，3-重甲）：", 1, 3)
    
    # 加载武器和子弹数据
    weapons = load_weapons_data()
    bullets = load_bullets_data()
    
    # 选择武器
    print("\n=== 选择武器 ===")
    print("（点射武器与多弹丸武器不适用本程序）")
    
    # 按类别分组武器
    categories = {}
    for weapon in weapons:
        if weapon['category'] not in categories:
            categories[weapon['category']] = []
        categories[weapon['category']].append(weapon)
    
    # 显示武器类别
    print("\n请选择武器类别：")
    category_list = list(categories.keys())
    for i, category in enumerate(category_list, 1):
        print(f"{i}. {category}")
    
    category_choice = get_int_input("输入类别编号：", 1, len(category_list))
    selected_category = category_list[category_choice - 1]
    
    # 显示选定类别中的武器
    print(f"\n{selected_category}武器列表：")
    category_weapons = categories[selected_category]
    for i, weapon in enumerate(category_weapons, 1):
        print(f"{i}. {weapon['name']}")
    
    weapon_choice = get_int_input("输入武器编号：", 1, len(category_weapons))
    selected_weapon = category_weapons[weapon_choice - 1]
    
    # 选择子弹
    print("\n=== 选择子弹 ===")
    # 按口径分组子弹
    calibers = {}
    for bullet in bullets:
        if bullet['caliber'] not in calibers:
            calibers[bullet['caliber']] = []
        calibers[bullet['caliber']].append(bullet)
    
    # 显示子弹口径
    print("\n请选择子弹口径：")
    caliber_list = list(calibers.keys())
    for i, caliber in enumerate(caliber_list, 1):
        print(f"{i}. {caliber}")
    
    caliber_choice = get_int_input("输入口径编号：", 1, len(caliber_list))
    selected_caliber = caliber_list[caliber_choice - 1]
    
    # 显示选定口径中的子弹
    print(f"\n{selected_caliber}子弹列表：")
    caliber_bullets = calibers[selected_caliber]
    for i, bullet in enumerate(caliber_bullets, 1):
        pen_level = bullet['penetration_level']
        dmg_mult = bullet['base_damage_multiplier']
        print(f"{i}. {bullet['name']} (穿透等级:{pen_level}, 伤害倍率:{dmg_mult})")
    
    bullet_choice = get_int_input("输入子弹编号：", 1, len(caliber_bullets))
    selected_bullet = caliber_bullets[bullet_choice - 1]
    
    # 输入目标距离
    distance = get_decimal_input("\n请输入目标距离（0-400米）：", 0.0, 400.0, 1)
    
    # 计算武器衰减倍率
    weapon_decay_multiplier = calculate_weapon_decay(float(distance), selected_weapon)
    
    # 获取护甲衰减倍率 (根据护甲等级)
    armor_decay_multiplier = Decimal(str(
        selected_bullet['armor_decay_factors'][armor_level - 1]
        if armor_level > 0 else 0.0
    ))
    
    # 设置武器参数
    weapon_damage = selected_weapon['base_damage']
    weapon_armor_damage = selected_weapon['armor_damage']
    fire_rate = selected_weapon['fire_rate']
    fire_mode = selected_weapon['fire_mode']
    trigger_delay = selected_weapon['trigger_delay']
    
    # 设置子弹参数
    penetration_level = selected_bullet['penetration_level']
    base_damage_multiplier = Decimal(str(selected_bullet['base_damage_multiplier']))
    base_armor_multiplier = Decimal(str(selected_bullet['base_armor_multiplier']))
    
    # 设置部位倍率
    body_part_multipliers = {
        '头部': Decimal(str(selected_weapon['head_multiplier'])),
        '胸部': Decimal(str(selected_weapon['chest_multiplier'])),
        '腹部': Decimal(str(selected_weapon['abdomen_multiplier'])),
        '大臂': Decimal(str(selected_weapon['upper_arm_multiplier'])),
        '小臂': Decimal(str(selected_weapon['lower_arm_multiplier'])),
        '大腿': Decimal(str(selected_weapon['thigh_multiplier'])),
        '小腿': Decimal(str(selected_weapon['calf_multiplier'])),
    }
    body_part_multipliers['下腹部'] = body_part_multipliers['腹部']
    
    # 打印选择的武器和子弹信息
    print("\n=== 武器信息 ===")
    print(f"武器: {selected_weapon['name']}")
    print(f"子弹: {selected_bullet['name']}")
    print(f"距离: {distance}米, 武器衰减倍率: {weapon_decay_multiplier}")
    print(f"穿透等级: {penetration_level}, 伤害倍率: {base_damage_multiplier}")
    print(f"护甲倍率: {base_armor_multiplier}, 护甲衰减倍率: {armor_decay_multiplier}")
    
    # 计算射击间隔
    shot_interval = (Decimal('60000') / Decimal(str(fire_rate))).quantize(
        Decimal('0.01'), rounding=ROUND_HALF_UP
    )
    
    # 初始化状态
    player_health = Decimal('100.0')
    current_helmet_durability = helmet_durability
    current_armor_durability = armor_durability
    protected_areas = {
        1: ['胸部', '腹部'],
        2: ['胸部', '腹部', '下腹部'],
        3: ['胸部', '腹部', '下腹部', '大臂']
    }[armor_type]
    valid_parts = ['头部', '胸部', '腹部', '大臂', '小臂', '大腿', '小腿', '下腹部', '未命中']
    
    hit_count = 0
    total_time = Decimal('0.0')
    total_damage = Decimal('0.0')
    total_armor_damage = Decimal('0.0')
    hit_statistics = {part: 0 for part in valid_parts}
    
    print("\n=== 开始模拟计算 ===")
    while True:
        # 输入命中部位
        while True:
            hit_part = input("\n请输入命中部位（或输入'未命中'/exit）：").strip()
            if hit_part.lower() == 'exit':
                return
            if hit_part in valid_parts:
                hit_statistics[hit_part] += 1
                hit_count += 1
                break
            print("无效输入，请重新输入。")
        
        # 计算总耗时
        trigger_delay_dec = Decimal(str(trigger_delay))
        if fire_mode == 1:  # 全自动
            total_time = trigger_delay_dec + shot_interval * Decimal(str(hit_count - 1))
        else:  # 半自动
            total_time = (trigger_delay_dec * Decimal(str(hit_count))) + (shot_interval * Decimal(str(hit_count - 1)))
        total_time = total_time.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        # 处理未命中
        if hit_part == '未命中':
            print("\n=== 计算结果 ===")
            print("本次攻击未命中")
            print(f"累计耗时：{total_time} ms")
            continue
        
        # 伤害计算逻辑
        is_protected = False
        protector_level = 0
        protector_type = None
        current_protector_durability = Decimal('0.0')
        
        # 判断保护状态
        if hit_part == '头部':
            if helmet_level > 0 and current_helmet_durability > Decimal('0'):
                is_protected = True
                protector_level = helmet_level
                protector_type = 'helmet'
                current_protector_durability = current_helmet_durability
        else:
            if armor_level > 0 and current_armor_durability > Decimal('0') and hit_part in protected_areas:
                is_protected = True
                protector_level = armor_level
                protector_type = 'armor'
                current_protector_durability = current_armor_durability
        
        # 计算穿透倍率
        penetration_multiplier = Decimal('0.0')
        if is_protected:
            diff = penetration_level - protector_level
            if diff < 0:
                penetration_multiplier = Decimal('0.0')
            elif diff == 0:
                penetration_multiplier = Decimal('0.5')
            elif diff == 1:
                penetration_multiplier = Decimal('0.75')
            else:
                penetration_multiplier = Decimal('1.0')
        
        # 计算护甲伤害
        weapon_armor_damage_dec = Decimal(str(weapon_armor_damage))
        armor_damage_value = weapon_armor_damage_dec * base_armor_multiplier * armor_decay_multiplier * weapon_decay_multiplier
        final_damage = Decimal('0.0')
        protector_destroyed = False
        armor_damage_dealt = Decimal('0.0')
        
        if is_protected:
            # 计算剩余耐久
            remaining_durability = current_protector_durability - armor_damage_value
            if remaining_durability <= Decimal('0'):
                protector_destroyed = True
                remaining_durability = Decimal('0.0')
            else:
                remaining_durability = remaining_durability.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
            
            # 累计护甲伤害
            armor_damage_dealt = current_protector_durability - remaining_durability
            total_armor_damage += armor_damage_dealt
            
            # 计算伤害
            part_multiplier = body_part_multipliers[hit_part]
            denominator = weapon_armor_damage_dec * base_armor_multiplier * armor_decay_multiplier * weapon_decay_multiplier
            
            if denominator == Decimal('0'):
                ratio = Decimal('0.0')
            else:
                ratio = current_protector_durability / denominator
            
            weapon_damage_dec = Decimal(str(weapon_damage))
            
            if current_protector_durability >= armor_damage_value:
                final_damage = weapon_damage_dec * base_damage_multiplier * part_multiplier * penetration_multiplier * weapon_decay_multiplier
            else:
                part1 = ratio * weapon_damage_dec * base_damage_multiplier * part_multiplier * penetration_multiplier * weapon_decay_multiplier
                part2 = (Decimal('1') - ratio) * weapon_damage_dec * base_damage_multiplier * part_multiplier * weapon_decay_multiplier
                final_damage = part1 + part2
            
            # 四舍五入伤害值
            final_damage = final_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            # 更新耐久
            if protector_type == 'helmet':
                current_helmet_durability = remaining_durability
            else:
                current_armor_durability = remaining_durability
        else:
            # 未受保护
            part_multiplier = body_part_multipliers[hit_part]
            weapon_damage_dec = Decimal(str(weapon_damage))
            final_damage = weapon_damage_dec * base_damage_multiplier * part_multiplier * weapon_decay_multiplier
            final_damage = final_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        total_damage += final_damage
        player_health -= final_damage
        
        # 输出结果
        print("\n=== 计算结果 ===")
        if is_protected:
            if protector_type == 'helmet':
                print("头盔被击碎！" if protector_destroyed else "头盔未被击碎。")
                print(f"头盔损失耐久：{armor_damage_dealt.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}")
            else:
                print("护甲被击碎！" if protector_destroyed else "护甲未被击碎。")
                print(f"护甲损失耐久：{armor_damage_dealt.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}")
        elif hit_part == '头部' and helmet_level > 0:
            print("（未受头盔保护）")
        
        print(f"受到伤害：{final_damage}")
        print(f"剩余生命值：{player_health.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}")
        print(f"剩余头盔耐久：{current_helmet_durability}")
        print(f"剩余护甲耐久：{current_armor_durability}")
        print(f"累计耗时：{total_time} ms")
        
        # 死亡处理
        if player_health <= Decimal('0'):
            print("\n=== 最终统计 ===")
            print(f"射击模式：{'全自动' if fire_mode == 1 else '半自动'}")
            print(f"总造成伤害：{total_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}")
            print(f"总护甲伤害：{total_armor_damage.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}")
            
            print("\n命中统计：")
            total_shots = sum(hit_statistics.values())
            valid_hits = total_shots - hit_statistics['未命中']
            display_order = ['未命中', '头部', '胸部', '腹部', '下腹部', '大臂', '小臂', '大腿', '小腿']
            for part in display_order:
                count = hit_statistics.get(part, 0)
                if count > 0:
                    print(f"{part.ljust(5)}：{count}次")
            
            print(f"\n有效命中次数：{valid_hits}次")
            print(f"总攻击次数：{total_shots}次")
            print(f"击杀耗时：{total_time} ms")
            
            input("\n按回车键结束模拟计算...")
            break

if __name__ == "__main__":
    main()