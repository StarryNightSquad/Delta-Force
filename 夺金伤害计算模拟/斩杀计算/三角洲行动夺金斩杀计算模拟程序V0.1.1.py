import openpyxl
from decimal import Decimal, ROUND_HALF_UP, ROUND_FLOOR, ROUND_CEILING
import msvcrt  # 用于检测按键
import sys  # 用于读取命令行参数

# 全局调试模式标志
DEBUG_MODE = False

def parse_command_line_args():
    """解析命令行参数"""
    global DEBUG_MODE
    if '--debug' in sys.argv:
        DEBUG_MODE = True
        print("调试模式已启用")

def debug_print(message):
    """调试信息输出"""
    if DEBUG_MODE:
        print(f"[DEBUG] {message}")

def load_weapon_data():
    """加载武器数据"""
    try:
        wb = openpyxl.load_workbook('S6夺金武器.xlsx', data_only=True)
        ws = wb['夺金模式']
        
        weapons = []
        current_category = ""
        
        for row in range(3, ws.max_row + 1):
            category_cell = ws.cell(row=row, column=1).value
            if category_cell and str(category_cell).strip():
                current_category = str(category_cell).strip()
                continue
                
            weapon_name = ws.cell(row=row, column=2).value
            if not weapon_name or not str(weapon_name).strip():
                continue
                
            weapons.append({
                'category': current_category,
                'name': str(weapon_name).strip(),
                'caliber': ws.cell(row=row, column=3).value,
                'base_damage': ws.cell(row=row, column=9).value or 0,
                'armor_damage': ws.cell(row=row, column=11).value or 0,
                'head_multiplier': ws.cell(row=row, column=13).value or 1.0
            })
        
        return weapons
    
    except Exception as e:
        print(f"无法加载武器数据: {e}")
        return []

def load_bullet_data():
    """加载子弹数据"""
    try:
        wb = openpyxl.load_workbook('S6子弹数据.xlsx', data_only=True)
        ws = wb['子弹数据']
        
        bullets = []
        current_caliber = ""
        
        for row in range(3, ws.max_row + 1):
            caliber_cell = ws.cell(row=row, column=1).value
            if caliber_cell and str(caliber_cell).strip():
                current_caliber = str(caliber_cell).strip()
                
            bullet_name = ws.cell(row=row, column=2).value
            if not bullet_name or not str(bullet_name).strip():
                continue
                
            # 获取同级穿透倍率（N列）
            same_level_penetration = ws.cell(row=row, column=14).value
            if same_level_penetration is None or same_level_penetration == "":
                same_level_penetration = 0.5  # 默认值
                
            bullets.append({
                'caliber': current_caliber,
                'name': str(bullet_name).strip(),
                'penetration_level': ws.cell(row=row, column=4).value or 0,
                'base_damage_multiplier': ws.cell(row=row, column=6).value or 1.0,
                'base_armor_multiplier': ws.cell(row=row, column=7).value or 1.0,
                'same_level_penetration': same_level_penetration,
                'armor_decay_factors': [
                    ws.cell(row=row, column=8).value or 0.0,  # 1甲
                    ws.cell(row=row, column=9).value or 0.0,  # 2甲
                    ws.cell(row=row, column=10).value or 0.0, # 3甲
                    ws.cell(row=row, column=11).value or 0.0, # 4甲
                    ws.cell(row=row, column=12).value or 0.0, # 5甲
                    ws.cell(row=row, column=13).value or 0.0  # 6甲
                ]
            })
        
        return bullets
    
    except Exception as e:
        print(f"无法加载子弹数据: {e}")
        return []

def calculate_damage(weapon, bullet, armor_level, durability=None, double_shot=False):
    """计算伤害"""
    # 转换为Decimal以确保精度
    base_damage = Decimal(str(weapon['base_damage']))
    weapon_armor_damage = Decimal(str(weapon['armor_damage']))
    head_multiplier = Decimal(str(weapon['head_multiplier']))
    base_damage_multiplier = Decimal(str(bullet['base_damage_multiplier']))
    base_armor_multiplier = Decimal(str(bullet['base_armor_multiplier']))
    
    # 使用子弹数据表格的N列数据（同级穿透倍率）
    penetration_multiplier = Decimal(str(bullet['same_level_penetration']))
    
    # 获取护甲衰减倍率
    armor_decay_factor = Decimal(str(bullet['armor_decay_factors'][armor_level - 1]))
    
    # 计算护甲伤害值
    armor_damage_value = (weapon_armor_damage * base_armor_multiplier * 
                         armor_decay_factor * Decimal('1.0'))
    
    # 如果是双发，伤害和护甲伤害都翻倍
    if double_shot:
        base_damage *= 2
        armor_damage_value *= 2
    
    # 计算无护甲时的爆头伤害
    unarmored_damage = base_damage * base_damage_multiplier * head_multiplier * Decimal('1.0')
    
    # 计算有护甲且足够时的爆头伤害
    armored_damage = base_damage * base_damage_multiplier * head_multiplier * penetration_multiplier * Decimal('1.0')
    
    # 输出调试信息
    debug_print(f"基础伤害: {base_damage}")
    debug_print(f"武器护甲伤害: {weapon_armor_damage}")
    debug_print(f"爆头倍率: {head_multiplier}")
    debug_print(f"子弹伤害倍率: {base_damage_multiplier}")
    debug_print(f"子弹护甲倍率: {base_armor_multiplier}")
    debug_print(f"同级穿透倍率: {penetration_multiplier}")
    debug_print(f"护甲衰减倍率: {armor_decay_factor}")
    debug_print(f"护甲伤害值: {armor_damage_value}")
    debug_print(f"无护甲伤害: {unarmored_damage}")
    debug_print(f"有护甲伤害: {armored_damage}")
    
    # 如果没有提供耐久值，直接返回伤害值
    if durability is None:
        return unarmored_damage, armored_damage, armor_damage_value
    
    # 计算实际伤害
    durability_dec = Decimal(str(durability))
    
    # 如果耐久足够承受全部护甲伤害
    if durability_dec >= armor_damage_value:
        final_damage = armored_damage
        debug_print(f"耐久足够，最终伤害: {final_damage}")
    else:
        # 计算比例
        ratio = durability_dec / armor_damage_value
        # 计算混合伤害
        part1 = ratio * armored_damage
        part2 = (Decimal('1') - ratio) * unarmored_damage
        final_damage = part1 + part2
        debug_print(f"耐久不足，比例: {ratio}, 混合伤害: {final_damage} (部分穿透: {part1} + 无护甲: {part2})")
    
    return final_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def verify_min_durability(weapon, bullet, armor_level, min_durability, double_shot=False):
    """验证最小耐久值"""
    damage = calculate_damage(weapon, bullet, armor_level, float(min_durability), double_shot)
    debug_print(f"验证耐久 {min_durability} 时的伤害: {damage}")
    return damage < 100

def calculate_min_durability(weapon, bullet, armor_level):
    """计算最小耐久值"""
    debug_print(f"\n=== 计算 {weapon['name']} + {bullet['name']} 对 {armor_level}级护甲 ===")
    
    # 计算单发伤害
    unarmored_damage, armored_damage, armor_damage_value = calculate_damage(weapon, bullet, armor_level)
    
    debug_print(f"单发无护甲伤害: {unarmored_damage}")
    debug_print(f"单发有护甲伤害: {armored_damage}")
    debug_print(f"单发护甲伤害值: {armor_damage_value}")
    
    # 情况1: 单发爆头伤害经过护甲减伤后依然大于等于100
    if armored_damage >= 100:
        debug_print("情况1: 单发爆头伤害经过护甲减伤后依然大于等于100")
        return "单发必死"
    
    # 情况2: 单发爆头伤害大于100，但减伤后小于100
    if unarmored_damage >= 100:
        debug_print("情况2: 单发爆头伤害大于100，但减伤后小于100")
        # 解方程: unarmored_damage * (1 - ratio) + armored_damage * ratio < 100
        # 其中 ratio = durability / armor_damage_value
        # 化简得: durability > armor_damage_value * (unarmored_damage - 100) / (unarmored_damage - armored_damage)
        if unarmored_damage == armored_damage:
            debug_print("无解，单发必死")
            return "单发必死"  # 无解，单发必死
            
        ratio_threshold = (unarmored_damage - Decimal('100')) / (unarmored_damage - armored_damage)
        min_durability = ratio_threshold * armor_damage_value
        
        debug_print(f"比例阈值: {ratio_threshold}")
        debug_print(f"理论最小耐久: {min_durability}")
        
        # 计算去尾和进一值
        floor_durability = min_durability.quantize(Decimal('0.1'), rounding=ROUND_FLOOR)
        ceil_durability = min_durability.quantize(Decimal('0.1'), rounding=ROUND_CEILING)
        
        debug_print(f"去尾值: {floor_durability}")
        debug_print(f"进一值: {ceil_durability}")
        
        # 验证去尾值
        if verify_min_durability(weapon, bullet, armor_level, floor_durability):
            debug_print(f"去尾值 {floor_durability} 满足条件")
            return f"单发斩杀耐久: {float(floor_durability):.1f}"
        
        # 验证进一值
        if verify_min_durability(weapon, bullet, armor_level, ceil_durability):
            debug_print(f"进一值 {ceil_durability} 满足条件")
            return f"单发斩杀耐久: {float(ceil_durability):.1f}"
        
        debug_print("无解，单发必死")
        return "单发必死"  # 无解，单发必死
    
    # 情况3: 单发爆头伤害小于100，计算双发
    debug_print("情况3: 单发爆头伤害小于100，计算双发")
    double_unarmored, double_armored, double_armor_damage = calculate_damage(weapon, bullet, armor_level, None, True)
    
    debug_print(f"双发无护甲伤害: {double_unarmored}")
    debug_print(f"双发有护甲伤害: {double_armored}")
    debug_print(f"双发护甲伤害值: {double_armor_damage}")
    
    # 情况4: 双发无护甲伤害小于100
    if double_unarmored < 100:
        debug_print("情况4: 双发无护甲伤害小于100")
        return "双发无法斩杀"
    
    # 情况5: 双发有护甲伤害大于等于100
    if double_armored >= 100:
        debug_print("情况5: 双发有护甲伤害大于等于100")
        return "双发必死"
    
    # 情况6: 双发无护甲伤害大于等于100，但有护甲伤害小于100
    debug_print("情况6: 双发无护甲伤害大于等于100，但有护甲伤害小于100")
    # 计算双发不可击杀的最低耐久
    # 解方程: double_unarmored * (1 - ratio) + double_armored * ratio < 100
    # 其中 ratio = durability / double_armor_damage
    # 化简得: durability > double_armor_damage * (double_unarmored - 100) / (double_unarmored - double_armored)
    if double_unarmored == double_armored:
        debug_print("无解，双发必死")
        return "双发必死"  # 无解，双发必死
        
    ratio_threshold = (double_unarmored - Decimal('100')) / (double_unarmored - double_armored)
    min_durability = ratio_threshold * double_armor_damage
    
    debug_print(f"比例阈值: {ratio_threshold}")
    debug_print(f"理论最小耐久: {min_durability}")
    
    # 计算去尾和进一值
    floor_durability = min_durability.quantize(Decimal('0.1'), rounding=ROUND_FLOOR)
    ceil_durability = min_durability.quantize(Decimal('0.1'), rounding=ROUND_CEILING)
    
    debug_print(f"去尾值: {floor_durability}")
    debug_print(f"进一值: {ceil_durability}")
    
    # 验证去尾值
    if verify_min_durability(weapon, bullet, armor_level, floor_durability, True):
        debug_print(f"去尾值 {floor_durability} 满足条件")
        return f"双发斩杀耐久: {float(floor_durability):.1f}"
    
    # 验证进一值
    if verify_min_durability(weapon, bullet, armor_level, ceil_durability, True):
        debug_print(f"进一值 {ceil_durability} 满足条件")
        return f"双发斩杀耐久: {float(ceil_durability):.1f}"
    
    debug_print("无解，双发必死")
    return "双发必死"  # 无解，双发必死

def wait_for_continue():
    """等待用户按Enter继续或Esc退出"""
    print("\n按Enter键开始新一轮计算，或按Esc键退出...")
    while True:
        if msvcrt.kbhit():
            key = msvcrt.getch()
            if key == b'\r':  # Enter键
                return True
            elif key == b'\x1b':  # Esc键
                return False

def main():
    # 解析命令行参数
    parse_command_line_args()
    
    while True:
        print("三角洲行动夺金伤害计算模拟程序 - 爆头不可击杀最低耐久计算")
        
        # 加载数据
        weapons = load_weapon_data()
        bullets = load_bullet_data()
        
        if not weapons or not bullets:
            print("无法加载数据，请确保Excel文件存在且格式正确")
            return
        
        # 按武器分类分组
        categories = {}
        for weapon in weapons:
            if weapon['category'] not in categories:
                categories[weapon['category']] = []
            categories[weapon['category']].append(weapon)
        
        # 选择武器分类
        print("\n可用武器分类:")
        category_list = list(categories.keys())
        for i, category in enumerate(category_list, 1):
            print(f"{i}. {category}")
        
        try:
            category_choice = int(input("\n请选择武器分类 (输入编号): ")) - 1
            selected_category = category_list[category_choice]
            print(f"已选择分类: {selected_category}")
        except (ValueError, IndexError):
            print("无效选择")
            if not wait_for_continue():
                break
            continue
        
        # 选择具体武器
        category_weapons = categories[selected_category]
        print(f"\n{selected_category}武器列表:")
        for i, weapon in enumerate(category_weapons, 1):
            print(f"{i}. {weapon['name']} ({weapon['caliber']})")
        
        try:
            weapon_choice = int(input("\n请选择武器 (输入编号): ")) - 1
            selected_weapon = category_weapons[weapon_choice]
        except (ValueError, IndexError):
            print("无效选择")
            if not wait_for_continue():
                break
            continue
        
        # 显示武器详细信息（只显示要求的字段）
        print(f"\n=== 武器详细信息 ===")
        print(f"名称: {selected_weapon['name']}")
        print(f"口径: {selected_weapon['caliber']}")
        print(f"基础伤害: {selected_weapon['base_damage']}")
        print(f"护甲伤害: {selected_weapon['armor_damage']}")
        print(f"爆头倍率: {selected_weapon['head_multiplier']}")
        
        # 筛选匹配口径的子弹
        caliber_bullets = [b for b in bullets if b['caliber'] == selected_weapon['caliber']]
        
        if not caliber_bullets:
            print(f"\n没有找到匹配 {selected_weapon['caliber']} 口径的子弹")
            if not wait_for_continue():
                break
            continue
        
        # 计算并显示每种子弹的结果
        print(f"\n=== {selected_weapon['name']} 爆头不可击杀最低耐久 ===")
        print("=" * 60)
        
        for bullet in caliber_bullets:
            results = []
            
            # 计算1-6级护甲的结果
            for armor_level in range(1, 7):
                # 只计算甲弹同级的情况
                if bullet['penetration_level'] != armor_level:
                    continue
                    
                result = calculate_min_durability(selected_weapon, bullet, armor_level)
                results.append(f"{armor_level}级: {result}")
            
            # 显示结果
            if results:
                print(f"{bullet['name']} (穿透等级{bullet['penetration_level']}, 同级穿透倍率{bullet['same_level_penetration']}): {', '.join(results)}")
            else:
                print(f"{bullet['name']}: 无同级护甲数据")
        
        # 询问是否继续
        if not wait_for_continue():
            break

if __name__ == "__main__":
    main()