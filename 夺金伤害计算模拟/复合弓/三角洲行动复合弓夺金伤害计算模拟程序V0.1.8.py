print("æœ¬ç¨‹åºç”±Bç«™ç¹æ˜Ÿæ”»ç•¥ç»„åˆ¶ä½œ")
print("æ³¨ï¼šå—é™äºæ•°æ®ç²¾åº¦é—®é¢˜ï¼Œæœ¬ç¨‹åºç»™å‡ºçš„æ‰€æœ‰æ—¶é—´ç›¸å…³è®¡ç®—ä»…ä¾›å‚è€ƒï¼Œä¸å®é™…å­˜åœ¨ä¸€å®šè¯¯å·®")

from decimal import Decimal, ROUND_HALF_UP
import openpyxl  # æ·»åŠ Excelå¤„ç†åº“

# ä½¿ç”¨Decimalè¿›è¡Œé«˜ç²¾åº¦è®¡ç®—
def round_decimal(value, decimals):
    if isinstance(value, Decimal):
        d = value
    else:
        d = Decimal(str(value))
    rounded = d.quantize(Decimal('1.' + '0'*decimals), rounding=ROUND_HALF_UP)
    return rounded

# éƒ¨ä½å€ç‡å­—å…¸ï¼ˆåŸºç¡€å€¼ï¼‰
base_location_multipliers = {
    "å¤´éƒ¨": Decimal('1.7'),
    "èƒ¸éƒ¨": Decimal('1'),
    "è…¹éƒ¨": Decimal('0.9'),
    "ä¸‹è…¹éƒ¨": Decimal('0.9'),
    "å¤§è‡‚": Decimal('0.45'),
    "å°è‡‚": Decimal('0.45'),
    "å¤§è…¿": Decimal('0.45'),
    "å°è…¿": Decimal('0.45')
}

# å¢å¼ºå¼“å¼¦åçš„éƒ¨ä½å€ç‡
enhanced_location_multipliers = {
    "å¤´éƒ¨": Decimal('2.0'),  # å¤´éƒ¨å€ç‡æå‡
    "èƒ¸éƒ¨": Decimal('1'),
    "è…¹éƒ¨": Decimal('0.9'),
    "ä¸‹è…¹éƒ¨": Decimal('0.9'),
    "å¤§è‡‚": Decimal('0.45'),
    "å°è‡‚": Decimal('0.45'),
    "å¤§è…¿": Decimal('0.45'),
    "å°è…¿": Decimal('0.45')
}

# ç®­çŸ¢å‚æ•°
arrows = {
    1: {
        "name": "ç»çº¤æŸ³å¶ç®­çŸ¢",
        "penetration": 3,
        "damage_multiplier": Decimal('1.0'),
        "armor_attenuation": [
            Decimal('0.9'), Decimal('0.9'), Decimal('0.9'), 
            Decimal('1.0'), Decimal('0.5'), Decimal('0.4')
        ],
        "armor_multiplier": Decimal('1.0')
    },
    2: {
        "name": "ç¢³çº¤ç»´åˆºéª¨ç®­çŸ¢",
        "penetration": 4,
        "damage_multiplier": Decimal('1.0'),
        "armor_attenuation": [
            Decimal('1.0'), Decimal('1.0'), Decimal('1.0'), 
            Decimal('1.0'), Decimal('1.1'), Decimal('0.6')
        ],
        "armor_multiplier": Decimal('1.0')
    },
    3: {
        "name": "ç¢³çº¤ç»´ç©¿ç”²ç®­çŸ¢",
        "penetration": 5,
        "damage_multiplier": Decimal('1.0'),
        "armor_attenuation": Decimal('1.1'),  # æ’å®šå€¼
        "armor_multiplier": Decimal('1.0')
    }
}

def get_decimal_input(prompt, min_val, max_val, decimals=1):
    while True:
        try:
            value_str = input(prompt)
            value = Decimal(value_str)
            value = round_decimal(value, decimals)
            if min_val <= value <= max_val:
                return value
            print(f"è¾“å…¥å€¼å¿…é¡»åœ¨ {min_val} åˆ° {max_val} ä¹‹é—´")
        except Exception:
            print("è¯·è¾“å…¥æœ‰æ•ˆçš„æ•°å­—")

def get_int_input(prompt, min_val, max_val):
    while True:
        try:
            value = int(input(prompt))
            if min_val <= value <= max_val:
                return value
            print(f"è¾“å…¥å€¼å¿…é¡»åœ¨ {min_val} åˆ° {max_val} ä¹‹é—´")
        except ValueError:
            print("è¯·è¾“å…¥æœ‰æ•ˆçš„æ•´æ•°")

def get_yes_no_input(prompt):
    while True:
        response = input(prompt).strip().lower()
        if response in ["y", "yes", "æ˜¯", "1"]:
            return True
        elif response in ["n", "no", "å¦", "0"]:
            return False
        print("è¯·è¾“å…¥ 'æ˜¯' æˆ– 'å¦'")

def get_arrow_armor_attenuation(arrow_type, armor_level):
    arrow_data = arrows[arrow_type]
    if arrow_type == 3:
        return arrow_data["armor_attenuation"]
    if 1 <= armor_level <= 6:
        return arrow_data["armor_attenuation"][armor_level - 1]
    return Decimal('1.0')  # é»˜è®¤å€¼

def calculate_penetration_rate(penetration, armor_level, is_head=False):
    """
    è®¡ç®—ç©¿é€ä¼¤å®³å€ç‡
    is_head: æ˜¯å¦å‘½ä¸­å¤´éƒ¨ï¼ˆå¤´éƒ¨ä½¿ç”¨ä¸åŒçš„è®¡ç®—è§„åˆ™ï¼‰
    """
    if penetration < armor_level:
        # ç©¿é€ç­‰çº§å°äºé˜²æŠ¤ç­‰çº§
        rate = Decimal('0.16') - Decimal(str(armor_level - penetration)) * Decimal('0.02')
        return max(rate, Decimal('0.02'))
    elif penetration == armor_level:
        # ç©¿é€ç­‰çº§ç­‰äºé˜²æŠ¤ç­‰çº§
        return Decimal('0.5') if is_head else Decimal('0.75')
    elif penetration == armor_level + 1:
        # ç©¿é€ç­‰çº§ç­‰äºé˜²æŠ¤ç­‰çº§+1
        return Decimal('0.75') if is_head else Decimal('0.9')
    else:  # penetration >= armor_level + 2
        # ç©¿é€ç­‰çº§å¤§äºç­‰äºé˜²æŠ¤ç­‰çº§+2
        return Decimal('1.0')

# === æ–°å¢çš„æŠ¤ç”²æ•°æ®åŠ è½½å’Œé€‰æ‹©å‡½æ•° ===
def load_armor_data():
    """åŠ è½½æŠ¤ç”²å’Œå¤´ç›”æ•°æ®"""
    try:
        wb = openpyxl.load_workbook('S5æŠ¤ç”²æ•°æ®.xlsx', data_only=True)
        ws = wb.active
        
        armors = []
        helmets = []
        current_section = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            
            # æ£€æµ‹ç« èŠ‚å¼€å§‹
            if cell_value == "æŠ¤ç”²":
                current_section = "armor"
                continue
            elif cell_value == "å¤´ç›”":
                current_section = "helmet"
                continue
                
            if not cell_value or not str(cell_value).strip():
                continue
                
            # è¯»å–æ•°æ®
            name = str(cell_value).strip()
            level = ws.cell(row=row, column=2).value
            armor_type = ws.cell(row=row, column=3).value
            max_durability = ws.cell(row=row, column=7).value  # Gåˆ—: åˆå§‹ä¸Šé™
            
            if level is None or max_durability is None:
                continue
                
            try:
                level = int(level)
                max_durability = Decimal(str(max_durability))
            except (ValueError, TypeError):
                continue
                
            item_data = {
                'name': name,
                'level': level,
                'max_durability': max_durability
            }
            
            if current_section == "armor":
                # æŠ¤ç”²ç±»å‹æ˜ å°„: åŠç”²->1, å…¨ç”²->2, é‡ç”²->3
                if armor_type == 'åŠç”²':
                    item_data['armor_type'] = 1
                elif armor_type == 'å…¨ç”²':
                    item_data['armor_type'] = 2
                elif armor_type == 'é‡ç”²':
                    item_data['armor_type'] = 3
                else:
                    continue  # è·³è¿‡æ— æ•ˆç±»å‹
                armors.append(item_data)
                
            elif current_section == "helmet":
                helmets.append(item_data)
                
        return armors, helmets
    
    except Exception as e:
        print(f"\né”™è¯¯: æ— æ³•åŠ è½½æŠ¤ç”²æ•°æ® - {e}")
        return [], []

def select_protection(items, item_type):
    """é€‰æ‹©é˜²æŠ¤è£…å¤‡å¹¶è¾“å…¥è€ä¹…"""
    print(f"\n=== é€‰æ‹©{item_type} ===")
    print("0. æ— ")
    
    # æŒ‰ç­‰çº§åˆ†ç»„
    levels = {}
    for item in items:
        level = item['level']
        if level not in levels:
            levels[level] = []
        levels[level].append(item)
    
    # æ˜¾ç¤ºç­‰çº§é€‰é¡¹
    sorted_levels = sorted(levels.keys())
    for i, level in enumerate(sorted_levels, 1):
        print(f"{i}. {level}çº§{item_type}")
    
    # è·å–ç”¨æˆ·é€‰æ‹©ï¼ŒåŒ…æ‹¬0ï¼ˆæ— ï¼‰é€‰é¡¹
    max_choice = len(sorted_levels)
    level_choice = get_int_input(f"è¯·é€‰æ‹©{item_type}ç­‰çº§ (0-{max_choice}): ", 0, max_choice)
    
    # å¤„ç†"æ— "é€‰é¡¹
    if level_choice == 0:
        return None, Decimal('0.0')
    
    selected_level = sorted_levels[level_choice - 1]
    
    # æ˜¾ç¤ºè¯¥ç­‰çº§ä¸‹çš„è£…å¤‡
    level_items = levels[selected_level]
    print(f"\n{selected_level}çº§{item_type}åˆ—è¡¨:")
    for i, item in enumerate(level_items, 1):
        print(f"{i}. {item['name']} (æœ€å¤§è€ä¹…: {item['max_durability']})")
    
    item_choice = get_int_input(f"è¯·é€‰æ‹©{item_type}: ", 1, len(level_items))
    selected_item = level_items[item_choice - 1]
    
    # è¾“å…¥å½“å‰è€ä¹…
    max_durability = float(selected_item['max_durability'])
    durability = get_decimal_input(
        f"è¯·è¾“å…¥å½“å‰{item_type}è€ä¹… (0.0-{max_durability}): ",
        0.0, max_durability, 1
    )
    
    return selected_item, durability

def main():
    print("===== å¤åˆå¼“ä¼¤å®³æ¨¡æ‹Ÿå™¨ =====")
    
    # === ä¿®æ”¹åçš„æŠ¤ç”²å’Œå¤´ç›”é€‰æ‹©éƒ¨åˆ† ===
    # åŠ è½½æŠ¤ç”²æ•°æ®
    print("æ­£åœ¨åŠ è½½æŠ¤ç”²æ•°æ®...")
    armors, helmets = load_armor_data()
    
    # åˆå§‹åŒ–æŠ¤ç”²å’Œå¤´ç›”ä¿¡æ¯
    helmet_level = 0
    helmet_durability = Decimal('0.0')
    armor_level = 0
    armor_durability = Decimal('0.0')
    armor_type = 0
    
    # é€‰æ‹©å¤´ç›”
    if helmets:
        selected_helmet, helmet_durability = select_protection(helmets, "å¤´ç›”")
        if selected_helmet:
            helmet_level = selected_helmet['level']
            print(f"å·²é€‰æ‹©å¤´ç›”: {selected_helmet['name']} (ç­‰çº§{helmet_level})")
        else:
            print("å·²é€‰æ‹©ï¼šæ— å¤´ç›”")
    else:
        print("æœªæ‰¾åˆ°å¤´ç›”æ•°æ®ï¼Œå°†ä½¿ç”¨æ— å¤´ç›”è®¾ç½®")
    
    # é€‰æ‹©æŠ¤ç”²
    if armors:
        selected_armor, armor_durability = select_protection(armors, "æŠ¤ç”²")
        if selected_armor:
            armor_level = selected_armor['level']
            armor_type = selected_armor['armor_type']
            print(f"å·²é€‰æ‹©æŠ¤ç”²: {selected_armor['name']} (ç­‰çº§{armor_level}, ç±»å‹{armor_type})")
        else:
            print("å·²é€‰æ‹©ï¼šæ— æŠ¤ç”²")
    else:
        print("æœªæ‰¾åˆ°æŠ¤ç”²æ•°æ®ï¼Œå°†ä½¿ç”¨æ— æŠ¤ç”²è®¾ç½®")
    
    # è®¾ç½®ä¿æŠ¤çš„èº«ä½“éƒ¨ä½
    protected_areas = {
        1: ["èƒ¸éƒ¨", "è…¹éƒ¨"],
        2: ["èƒ¸éƒ¨", "è…¹éƒ¨", "ä¸‹è…¹éƒ¨"],
        3: ["èƒ¸éƒ¨", "è…¹éƒ¨", "ä¸‹è…¹éƒ¨", "å¤§è‡‚"]
    }.get(armor_type, [])  # å¦‚æœæ²¡æœ‰æŠ¤ç”²ï¼Œè¿”å›ç©ºåˆ—è¡¨
    
    # å…¶ä»–è¾“å…¥ä¿æŒä¸å˜
    distance = get_decimal_input("è¾“å…¥ç›®æ ‡è·ç¦» (0.0-200.0): ", Decimal('0.0'), Decimal('200.0'))
    arrow_type = get_int_input("è¾“å…¥ç®­çŸ¢ç±»å‹ (1-ç»çº¤æŸ³å¶, 2-ç¢³çº¤ç»´åˆºéª¨, 3-ç¢³çº¤ç»´ç©¿ç”²): ", 1, 3)
    
    # æ˜¯å¦è£…é…å¢å¼ºå¼“å¼¦
    enhanced_string = get_yes_no_input("æ˜¯å¦è£…é…å¢å¼ºå¼“å¼¦? (æ˜¯/å¦): ")
    
    # æ ¹æ®æ˜¯å¦ä½¿ç”¨å¢å¼ºå¼“å¼¦é€‰æ‹©éƒ¨ä½å€ç‡å­—å…¸
    location_multipliers = enhanced_location_multipliers if enhanced_string else base_location_multipliers
    
    # æ­¦å™¨è¡°å‡å€ç‡ - å¢å¼ºå¼“å¼¦å½±å“è·ç¦»é˜ˆå€¼
    if enhanced_string:
        weapon_attenuation = Decimal('1.0') if distance <= Decimal('84.5') else Decimal('0.9')
    else:
        weapon_attenuation = Decimal('1.0') if distance <= Decimal('65') else Decimal('0.9')
    
    shot_interval = Decimal('500')  # æ¯«ç§’ï¼ˆå°„å‡»é—´éš”ï¼‰
    
    # ç»Ÿè®¡ä¿¡æ¯
    total_damage = Decimal('0.0')
    total_armor_damage = Decimal('0.0')
    location_hits = {loc: 0 for loc in location_multipliers}
    location_hits["æœªå‘½ä¸­"] = 0
    total_hits = 0
    total_shots = 0
    total_time_ms = Decimal('0')  # æ€»è€—æ—¶ï¼ˆæ¯«ç§’ï¼‰
    shot_type_count = {"ç¬å‘": 0, "æ»¡è“„": 0}  # è®°å½•å°„å‡»æ–¹å¼ä½¿ç”¨æ¬¡æ•°
    
    # æ˜¾ç¤ºå½“å‰é…ç½®
    print("\n===== å½“å‰é…ç½® =====")
    print(f"å¤´ç›”: {helmet_durability:.1f}è€ä¹…, {helmet_level}çº§é˜²æŠ¤")
    
    # æ˜¾ç¤ºæŠ¤ç”²ç±»å‹ä¸ºæ–‡å­—
    armor_type_str = {1: "åŠç”²", 2: "å…¨ç”²", 3: "é‡ç”²"}.get(armor_type, "æ— ")
    print(f"æŠ¤ç”²: {armor_durability:.1f}è€ä¹…, {armor_level}çº§é˜²æŠ¤, ç±»å‹: {armor_type_str}")
    
    print(f"è·ç¦»: {distance:.1f}ç±³")
    print(f"ç®­çŸ¢: {arrows[arrow_type]['name']}")
    print(f"å¢å¼ºå¼“å¼¦: {'å·²è£…é…' if enhanced_string else 'æœªè£…é…'}")
    print(f"æ­¦å™¨è¡°å‡å€ç‡: {weapon_attenuation:.1f} (è·ç¦»é˜ˆå€¼: {'84.5' if enhanced_string else '65'}ç±³)")
    if enhanced_string:
        print("å¤´éƒ¨å€ç‡æå‡è‡³2.0")
    print("="*20)
    print("æ¯æ¬¡æ”»å‡»å‰éœ€è¦é€‰æ‹©å°„å‡»æ–¹å¼ï¼š")
    print("1 = ç¬å‘ (åŸºç¡€ä¼¤å®³76.5, æŠ¤ç”²ä¼¤å®³55.25, æ‹‰å¼“æ—¶é—´80ms)")
    print("2 = æ»¡è“„ (åŸºç¡€ä¼¤å®³90.0, æŠ¤ç”²ä¼¤å®³65.0, æ‹‰å¼“æ—¶é—´{}ms)".format(460 if enhanced_string else 540))
    
    # ä¸»å¾ªç¯ï¼ˆä¿æŒä¸å˜ï¼‰
    while health > 0:
        # è¾“å…¥æœ¬æ¬¡å°„å‡»æ–¹å¼
        shot_mode = ""
        while shot_mode not in ["1", "2"]:
            shot_mode = input("\nè¾“å…¥æœ¬æ¬¡å°„å‡»æ–¹å¼ (1=ç¬å‘, 2=æ»¡è“„): ")
        
        if shot_mode == "1":
            # ç¬å‘
            base_damage_this = Decimal('76.5')
            base_armor_damage_this = Decimal('55.25')
            draw_time_this = Decimal('80')  # æ¯«ç§’ï¼ˆæ‹‰å¼“æ—¶é—´ï¼‰
            shot_type = "ç¬å‘"
        else:
            # æ»¡è“„
            base_damage_this = Decimal('90.0')
            base_armor_damage_this = Decimal('65.0')
            draw_time_this = Decimal('460') if enhanced_string else Decimal('540')  # æ¯«ç§’ï¼ˆæ‹‰å¼“æ—¶é—´ï¼‰
            shot_type = "æ»¡è“„"
        
        shot_type_count[shot_type] += 1
        
        location = input("è¾“å…¥å‘½ä¸­éƒ¨ä½ (å¤´éƒ¨/èƒ¸éƒ¨/è…¹éƒ¨/ä¸‹è…¹éƒ¨/å¤§è‡‚/å°è‡‚/å¤§è…¿/å°è…¿/æœªå‘½ä¸­)ï¼š ")
        
        # å¤„ç†æœªå‘½ä¸­
        if location == "æœªå‘½ä¸­":
            total_shots += 1
            location_hits["æœªå‘½ä¸­"] += 1
            # æ›´æ–°æ€»æ—¶é—´ï¼šæœ¬æ¬¡æ‹‰å¼“æ—¶é—´ + å¦‚æœæ˜¯ç¬¬ä¸€æ¬¡å°„å‡»åˆ™ä¸åŠ é—´éš”ï¼Œå¦åˆ™åŠ ä¸Šé—´éš”
            total_time_ms += draw_time_this
            if total_shots > 1:
                total_time_ms += shot_interval
            print(f"{shot_type}æœªå‘½ä¸­! è€—æ—¶: {float(total_time_ms/Decimal('1000')):.3f}ç§’")
            continue
        
        # éªŒè¯éƒ¨ä½è¾“å…¥
        if location not in location_multipliers:
            print("æ— æ•ˆéƒ¨ä½! è¯·é‡æ–°è¾“å…¥")
            continue
        
        total_shots += 1
        total_hits += 1
        location_hits[location] += 1
        
        # æ›´æ–°æ€»æ—¶é—´ï¼šæœ¬æ¬¡æ‹‰å¼“æ—¶é—´ + å¦‚æœæ˜¯ç¬¬ä¸€æ¬¡å°„å‡»åˆ™ä¸åŠ é—´éš”ï¼Œå¦åˆ™åŠ ä¸Šé—´éš”
        total_time_ms += draw_time_this
        if total_shots > 1:
            total_time_ms += shot_interval
        
        # è·å–éƒ¨ä½å€ç‡
        location_mult = location_multipliers[location]
        
        # ç¡®å®šä¿æŠ¤çŠ¶æ€
        protected = False
        armor_damage_taken = Decimal('0.0')
        penetration_rate = Decimal('1.0')
        armor_dmg_value = Decimal('0.0')
        gear_type = ""
        gear_durability = Decimal('0.0')
        gear_level = 0
        
        # å¤´éƒ¨å‘½ä¸­å¤„ç†
        if location == "å¤´éƒ¨":
            if helmet_level > 0 and helmet_durability > Decimal('0'):
                protected = True
                gear_type = "å¤´ç›”"
                gear_durability = helmet_durability
                gear_level = helmet_level
        # èº«ä½“éƒ¨ä½å‘½ä¸­å¤„ç†
        else:
            if armor_level > 0 and armor_durability > Decimal('0') and location in protected_areas:
                protected = True
                gear_type = "æŠ¤ç”²"
                gear_durability = armor_durability
                gear_level = armor_level
        
        # ä¼¤å®³è®¡ç®—
        if not protected or gear_level == 0:  # æœªå—ä¿æŠ¤
            damage = base_damage_this * location_mult * weapon_attenuation
            armor_damage_taken = Decimal('0.0')
        else:
            # è·å–ç®­çŸ¢å‚æ•°
            arrow_data = arrows[arrow_type]
            penetration = arrow_data["penetration"]
            
            # è®¡ç®—æŠ¤ç”²è¡°å‡å€ç‡
            armor_attenuation = get_arrow_armor_attenuation(arrow_type, gear_level)
            
            # è®¡ç®—æŠ¤ç”²ä¼¤å®³å€¼
            armor_dmg_value = base_armor_damage_this * arrow_data["armor_multiplier"] * armor_attenuation * weapon_attenuation
            armor_dmg_value = round_decimal(armor_dmg_value, 1)
            
            # è®¡ç®—ç©¿é€å€ç‡ - å¤´éƒ¨ä½¿ç”¨ç‰¹æ®Šè§„åˆ™
            is_head = (location == "å¤´éƒ¨")
            penetration_rate = calculate_penetration_rate(penetration, gear_level, is_head)
            
            # è®¡ç®—å®é™…ä¼¤å®³
            if gear_durability >= armor_dmg_value:
                # å…¬å¼a: æŠ¤ç”²å®Œå…¨å¸æ”¶
                damage = base_damage_this * arrow_data["damage_multiplier"] * location_mult * penetration_rate * weapon_attenuation
                armor_damage_taken = armor_dmg_value
                new_durability = gear_durability - armor_damage_taken
            else:
                # å…¬å¼b: æŠ¤ç”²éƒ¨åˆ†å‡»ç©¿
                ratio = gear_durability / armor_dmg_value
                damage_part1 = base_damage_this * arrow_data["damage_multiplier"] * location_mult * penetration_rate * weapon_attenuation
                damage_part2 = base_damage_this * arrow_data["damage_multiplier"] * location_mult * weapon_attenuation
                damage = ratio * damage_part1 + (1 - ratio) * damage_part2
                armor_damage_taken = gear_durability
                new_durability = Decimal('0.0')
            
            # æ›´æ–°è£…å¤‡è€ä¹…
            new_durability = round_decimal(new_durability, 1)  # è€ä¹…ç²¾ç¡®åˆ°0.1
            if gear_type == "å¤´ç›”":
                helmet_durability = max(new_durability, Decimal('0.0'))
            else:
                armor_durability = max(new_durability, Decimal('0.0'))
        
        # å››èˆäº”å…¥ä¼¤å®³å€¼
        damage = round_decimal(damage, 2)
        
        # æ›´æ–°ç©å®¶çŠ¶æ€
        health -= damage
        total_damage += damage
        total_armor_damage += armor_damage_taken
        
        # è¾“å‡ºç»“æœ
        print(f"\n{shot_type}å‘½ä¸­ {location}!")
        print(f"é€ æˆä¼¤å®³: {float(damage):.2f}")
        
        if protected:
            armor_damage_taken = round_decimal(armor_damage_taken, 1)
            print(f"{gear_type}æŸå¤±è€ä¹…: {float(armor_damage_taken):.1f}")
            if gear_durability > Decimal('0') and armor_damage_taken >= gear_durability:
                print(f"âš”ï¸ {gear_type}å·²å‡»ç¢!")
        
        if health > Decimal('0'):
            health_rounded = round_decimal(health, 1)
            print(f"å‰©ä½™ç”Ÿå‘½: {float(health_rounded):.1f}")
            print(f"å¤´ç›”è€ä¹…: {float(helmet_durability):.1f}")
            print(f"æŠ¤ç”²è€ä¹…: {float(armor_durability):.1f}")
        else:
            print("ğŸ’€ ä½ æ­»äº†!")
            print(f"æœ€ç»ˆå¤´ç›”è€ä¹…: {float(helmet_durability):.1f}")
            print(f"æœ€ç»ˆæŠ¤ç”²è€ä¹…: {float(armor_durability):.1f}")
            input("æŒ‰å›è½¦é”®ç»“æŸæ¨¡æ‹Ÿè®¡ç®—")
            break
    
    # è¾“å‡ºç»Ÿè®¡ä¿¡æ¯
    print("\n===== æˆ˜æ–—ç»Ÿè®¡ =====")
    print(f"æ€»ä¼¤å®³: {float(total_damage):.2f}")
    print(f"æ€»æŠ¤ç”²ä¼¤å®³: {float(total_armor_damage):.1f}")
    print(f"æ€»æ”»å‡»æ¬¡æ•°: {total_shots}")
    print(f"æ€»å‘½ä¸­æ¬¡æ•°: {total_hits}")
    print(f"æ€»è€—æ—¶: {float(total_time_ms/Decimal('1000')):.3f}ç§’")
    
    print("\nå°„å‡»æ–¹å¼ç»Ÿè®¡:")
    for shot_type, count in shot_type_count.items():
        if count > 0:
            print(f"{shot_type}: {count}æ¬¡")
    
    print("\néƒ¨ä½å‘½ä¸­ç»Ÿè®¡:")
    for loc, count in location_hits.items():
        if count > 0:
            print(f"{loc}: {count}æ¬¡")

if __name__ == "__main__":
    # ç©å®¶åˆå§‹çŠ¶æ€
    health = Decimal('100.0')
    main()
