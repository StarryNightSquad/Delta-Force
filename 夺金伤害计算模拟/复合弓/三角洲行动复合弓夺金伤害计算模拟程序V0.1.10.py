print("本程序由B站繁星攻略组制作")
print("注：受限于数据精度问题，本程序给出的所有时间相关计算仅供参考，与实际存在一定误差")

from decimal import Decimal, ROUND_HALF_UP
import openpyxl  # 添加Excel处理库

# 使用Decimal进行高精度计算
def round_decimal(value, decimals):
    if isinstance(value, Decimal):
        d = value
    else:
        d = Decimal(str(value))
    rounded = d.quantize(Decimal('1.' + '0'*decimals), rounding=ROUND_HALF_UP)
    return rounded

# 部位倍率字典（基础值）
base_location_multipliers = {
    "头部": Decimal('1.7'),
    "胸部": Decimal('1'),
    "腹部": Decimal('0.9'),
    "下腹部": Decimal('0.9'),
    "大臂": Decimal('0.45'),
    "小臂": Decimal('0.45'),
    "大腿": Decimal('0.45'),
    "小腿": Decimal('0.45')
}

# 增强弓弦后的部位倍率
enhanced_location_multipliers = {
    "头部": Decimal('2.0'),  # 头部倍率提升
    "胸部": Decimal('1'),
    "腹部": Decimal('0.9'),
    "下腹部": Decimal('0.9'),
    "大臂": Decimal('0.45'),
    "小臂": Decimal('0.45'),
    "大腿": Decimal('0.45'),
    "小腿": Decimal('0.45')
}

# 箭矢参数
arrows = {
    1: {
        "name": "玻纤柳叶箭矢",
        "penetration": 3,
        "damage_multiplier": Decimal('1.0'),
        "armor_attenuation": [
            Decimal('0.9'), Decimal('0.9'), Decimal('0.9'), 
            Decimal('1.0'), Decimal('0.5'), Decimal('0.4')
        ],
        "armor_multiplier": Decimal('1.0')
    },
    2: {
        "name": "碳纤维刺骨箭矢",
        "penetration": 4,
        "damage_multiplier": Decimal('1.0'),
        "armor_attenuation": [
            Decimal('1.0'), Decimal('1.0'), Decimal('1.0'), 
            Decimal('1.0'), Decimal('1.1'), Decimal('0.6')
        ],
        "armor_multiplier": Decimal('1.0')
    },
    3: {
        "name": "碳纤维穿甲箭矢",
        "penetration": 5,
        "damage_multiplier": Decimal('1.0'),
        "armor_attenuation": Decimal('1.1'),  # 恒定值
        "armor_multiplier": Decimal('1.0')
    }
}

def get_decimal_input(prompt, min_val, max_val, decimals=1):
    while True:
        try:
            value_str = input(prompt)
            value = Decimal(value_str)
            value = round_decimal(value, decimals)
            if min_val <= value <= max_val:
                return value
            print(f"输入值必须在 {min_val} 到 {max_val} 之间")
        except Exception:
            print("请输入有效的数字")

def get_int_input(prompt, min_val, max_val):
    while True:
        try:
            value = int(input(prompt))
            if min_val <= value <= max_val:
                return value
            print(f"输入值必须在 {min_val} 到 {max_val} 之间")
        except ValueError:
            print("请输入有效的整数")

def get_yes_no_input(prompt):
    while True:
        response = input(prompt).strip().lower()
        if response in ["y", "yes", "是", "1"]:
            return True
        elif response in ["n", "no", "否", "0"]:
            return False
        print("请输入 '是' 或 '否'")

def get_arrow_armor_attenuation(arrow_type, armor_level):
    arrow_data = arrows[arrow_type]
    if arrow_type == 3:
        return arrow_data["armor_attenuation"]
    if 1 <= armor_level <= 6:
        return arrow_data["armor_attenuation"][armor_level - 1]
    return Decimal('1.0')  # 默认值

def calculate_penetration_rate(penetration, armor_level, is_head=False):
    """
    计算穿透伤害倍率
    is_head: 是否命中头部（头部使用不同的计算规则）
    """
    if penetration < armor_level:
        # 穿透等级小于防护等级
        rate = Decimal('0.16') - Decimal(str(armor_level - penetration)) * Decimal('0.02')
        return max(rate, Decimal('0.02'))
    elif penetration == armor_level:
        # 穿透等级等于防护等级
        return Decimal('0.5') if is_head else Decimal('0.75')
    elif penetration == armor_level + 1:
        # 穿透等级等于防护等级+1
        return Decimal('0.75') if is_head else Decimal('0.9')
    else:  # penetration >= armor_level + 2
        # 穿透等级大于等于防护等级+2
        return Decimal('1.0')

# === 新增的护甲数据加载和选择函数 ===
def load_armor_data():
    """加载护甲和头盔数据"""
    try:
        wb = openpyxl.load_workbook('S5护甲数据.xlsx', data_only=True)
        ws = wb.active
        
        armors = []
        helmets = []
        current_section = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            
            # 检测章节开始
            if cell_value == "护甲":
                current_section = "armor"
                continue
            elif cell_value == "头盔":
                current_section = "helmet"
                continue
                
            if not cell_value or not str(cell_value).strip():
                continue
                
            # 读取数据
            name = str(cell_value).strip()
            level = ws.cell(row=row, column=2).value
            armor_type = ws.cell(row=row, column=3).value
            max_durability = ws.cell(row=row, column=7).value  # G列: 初始上限
            
            if level is None or max_durability is None:
                continue
                
            try:
                level = int(level)
                max_durability = Decimal(str(max_durability))
            except (ValueError, TypeError):
                continue
                
            item_data = {
                'name': name,
                'level': level,
                'max_durability': max_durability
            }
            
            if current_section == "armor":
                # 护甲类型映射: 半甲->1, 全甲->2, 重甲->3
                if armor_type == '半甲':
                    item_data['armor_type'] = 1
                elif armor_type == '全甲':
                    item_data['armor_type'] = 2
                elif armor_type == '重甲':
                    item_data['armor_type'] = 3
                else:
                    continue  # 跳过无效类型
                armors.append(item_data)
                
            elif current_section == "helmet":
                helmets.append(item_data)
                
        return armors, helmets
    
    except Exception as e:
        print(f"\n错误: 无法加载护甲数据 - {e}")
        return [], []

def select_protection(items, item_type):
    """选择防护装备并输入耐久"""
    print(f"\n=== 选择{item_type} ===")
    print("0. 无")
    
    # 按等级分组
    levels = {}
    for item in items:
        level = item['level']
        if level not in levels:
            levels[level] = []
        levels[level].append(item)
    
    # 显示等级选项
    sorted_levels = sorted(levels.keys())
    for i, level in enumerate(sorted_levels, 1):
        print(f"{i}. {level}级{item_type}")
    
    # 获取用户选择，包括0（无）选项
    max_choice = len(sorted_levels)
    level_choice = get_int_input(f"请选择{item_type}等级 (0-{max_choice}): ", 0, max_choice)
    
    # 处理"无"选项
    if level_choice == 0:
        return None, Decimal('0.0')
    
    selected_level = sorted_levels[level_choice - 1]
    
    # 显示该等级下的装备
    level_items = levels[selected_level]
    print(f"\n{selected_level}级{item_type}列表:")
    for i, item in enumerate(level_items, 1):
        print(f"{i}. {item['name']} (最大耐久: {item['max_durability']})")
    
    item_choice = get_int_input(f"请选择{item_type}: ", 1, len(level_items))
    selected_item = level_items[item_choice - 1]
    
    # 输入当前耐久
    max_durability = float(selected_item['max_durability'])
    durability = get_decimal_input(
        f"请输入当前{item_type}耐久 (0.0-{max_durability}): ",
        0.0, max_durability, 1
    )
    
    return selected_item, durability

def main():
    print("===== 复合弓伤害模拟器 =====")
    
    # 玩家初始状态
    health = Decimal('100.0')
    
    # 加载护甲数据
    print("正在加载护甲数据...")
    armors, helmets = load_armor_data()
    
    # 初始化护甲和头盔信息
    helmet_level = 0
    helmet_durability = Decimal('0.0')
    armor_level = 0
    armor_durability = Decimal('0.0')
    armor_type = 0
    
    # 选择头盔
    if helmets:
        selected_helmet, helmet_durability = select_protection(helmets, "头盔")
        if selected_helmet:
            helmet_level = selected_helmet['level']
            print(f"已选择头盔: {selected_helmet['name']} (等级{helmet_level})")
        else:
            print("已选择：无头盔")
    else:
        print("未找到头盔数据，将使用无头盔设置")
    
    # 选择护甲
    if armors:
        selected_armor, armor_durability = select_protection(armors, "护甲")
        if selected_armor:
            armor_level = selected_armor['level']
            armor_type = selected_armor['armor_type']
            print(f"已选择护甲: {selected_armor['name']} (等级{armor_level}, 类型{armor_type})")
        else:
            print("已选择：无护甲")
    else:
        print("未找到护甲数据，将使用无护甲设置")
    
    # 设置保护的身体部位
    protected_areas = {
        1: ["胸部", "腹部"],
        2: ["胸部", "腹部", "下腹部"],
        3: ["胸部", "腹部", "下腹部", "大臂"]
    }.get(armor_type, [])  # 如果没有护甲，返回空列表
    
    # 其他输入保持不变
    distance = get_decimal_input("输入目标距离 (0.0-200.0): ", Decimal('0.0'), Decimal('200.0'))
    arrow_type = get_int_input("输入箭矢类型 (1-玻纤柳叶, 2-碳纤维刺骨, 3-碳纤维穿甲): ", 1, 3)
    
    # 是否装配增强弓弦
    enhanced_string = get_yes_no_input("是否装配增强弓弦? (是/否): ")
    
    # 根据是否使用增强弓弦选择部位倍率字典
    location_multipliers = enhanced_location_multipliers if enhanced_string else base_location_multipliers
    
    # 武器衰减倍率 - 增强弓弦影响距离阈值
    if enhanced_string:
        weapon_attenuation = Decimal('1.0') if distance <= Decimal('84.5') else Decimal('0.9')
    else:
        weapon_attenuation = Decimal('1.0') if distance <= Decimal('65') else Decimal('0.9')
    
    shot_interval = Decimal('500')  # 毫秒（射击间隔）
    
    # 统计信息
    total_damage = Decimal('0.0')
    total_armor_damage = Decimal('0.0')
    location_hits = {loc: 0 for loc in location_multipliers}
    location_hits["未命中"] = 0
    total_hits = 0
    total_shots = 0
    total_time_ms = Decimal('0')  # 总耗时（毫秒）
    shot_type_count = {"瞬发": 0, "半蓄": 0, "满蓄": 0}  # 记录射击方式使用次数
    
    # 显示当前配置
    print("\n===== 当前配置 =====")
    print(f"头盔: {helmet_durability:.1f}耐久, {helmet_level}级防护")
    
    # 显示护甲类型为文字
    armor_type_str = {1: "半甲", 2: "全甲", 3: "重甲"}.get(armor_type, "无")
    print(f"护甲: {armor_durability:.1f}耐久, {armor_level}级防护, 类型: {armor_type_str}")
    
    print(f"距离: {distance:.1f}米")
    print(f"箭矢: {arrows[arrow_type]['name']}")
    print(f"增强弓弦: {'已装配' if enhanced_string else '未装配'}")
    print(f"武器衰减倍率: {weapon_attenuation:.1f} (距离阈值: {'84.5' if enhanced_string else '65'}米)")
    if enhanced_string:
        print("头部倍率提升至2.0")
    print("="*20)
    print("每次攻击前需要选择射击方式：")
    print("1 = 瞬发 (基础伤害76.5, 护甲伤害55.25, 拉弓时间80ms)")
    print("2 = 半蓄 (基础伤害81.0, 护甲伤害58.5, 拉弓时间{}ms)".format(400 if enhanced_string else 480))
    print("3 = 满蓄 (基础伤害90.0, 护甲伤害65.0, 拉弓时间{}ms)".format(460 if enhanced_string else 540))
    
    # 主循环
    while health > 0:
        # 输入本次射击方式
        shot_mode = ""
        while shot_mode not in ["1", "2", "3"]:
            shot_mode = input("\n输入本次射击方式 (1=瞬发, 2=半蓄, 3=满蓄): ")
        
        if shot_mode == "1":
            # 瞬发
            base_damage_this = Decimal('76.5')
            base_armor_damage_this = Decimal('55.25')
            draw_time_this = Decimal('80')  # 毫秒（拉弓时间）
            shot_type = "瞬发"
        elif shot_mode == "2":
            # 半蓄
            base_damage_this = Decimal('81.0')  # 满蓄伤害的0.9倍
            base_armor_damage_this = Decimal('58.5')  # 满蓄护甲伤害的0.9倍
            draw_time_this = Decimal('400') if enhanced_string else Decimal('480')  # 毫秒（拉弓时间）
            shot_type = "半蓄"
        else:
            # 满蓄
            base_damage_this = Decimal('90.0')
            base_armor_damage_this = Decimal('65.0')
            draw_time_this = Decimal('460') if enhanced_string else Decimal('540')  # 毫秒（拉弓时间）
            shot_type = "满蓄"
        
        shot_type_count[shot_type] += 1
        
        location = input("输入命中部位 (头部/胸部/腹部/下腹部/大臂/小臂/大腿/小腿/未命中)： ")
        
        # 处理未命中
        if location == "未命中":
            total_shots += 1
            location_hits["未命中"] += 1
            # 更新总时间：本次拉弓时间 + 如果是第一次射击则不加间隔，否则加上间隔
            total_time_ms += draw_time_this
            if total_shots > 1:
                total_time_ms += shot_interval
            print(f"{shot_type}未命中! 耗时: {float(total_time_ms/Decimal('1000')):.3f}秒")
            continue
        
        # 验证部位输入
        if location not in location_multipliers:
            print("无效部位! 请重新输入")
            continue
        
        total_shots += 1
        total_hits += 1
        location_hits[location] += 1
        
        # 更新总时间：本次拉弓时间 + 如果是第一次射击则不加间隔，否则加上间隔
        total_time_ms += draw_time_this
        if total_shots > 1:
            total_time_ms += shot_interval
        
        # 获取部位倍率
        location_mult = location_multipliers[location]
        
        # 确定保护状态
        protected = False
        armor_damage_taken = Decimal('0.0')
        penetration_rate = Decimal('1.0')
        armor_dmg_value = Decimal('0.0')
        gear_type = ""
        gear_durability = Decimal('0.0')
        gear_level = 0
        
        # 头部命中处理
        if location == "头部":
            if helmet_level > 0 and helmet_durability > Decimal('0'):
                protected = True
                gear_type = "头盔"
                gear_durability = helmet_durability
                gear_level = helmet_level
        # 身体部位命中处理
        else:
            if armor_level > 0 and armor_durability > Decimal('0') and location in protected_areas:
                protected = True
                gear_type = "护甲"
                gear_durability = armor_durability
                gear_level = armor_level
        
        # 伤害计算
        if not protected or gear_level == 0:  # 未受保护
            damage = base_damage_this * location_mult * weapon_attenuation
            armor_damage_taken = Decimal('0.0')
        else:
            # 获取箭矢参数
            arrow_data = arrows[arrow_type]
            penetration = arrow_data["penetration"]
            
            # 计算护甲衰减倍率
            armor_attenuation = get_arrow_armor_attenuation(arrow_type, gear_level)
            
            # 计算护甲伤害值
            armor_dmg_value = base_armor_damage_this * arrow_data["armor_multiplier"] * armor_attenuation * weapon_attenuation
            armor_dmg_value = round_decimal(armor_dmg_value, 1)
            
            # 计算穿透倍率 - 头部使用特殊规则
            is_head = (location == "头部")
            penetration_rate = calculate_penetration_rate(penetration, gear_level, is_head)
            
            # 计算实际伤害
            if gear_durability >= armor_dmg_value:
                # 公式a: 护甲完全吸收
                damage = base_damage_this * arrow_data["damage_multiplier"] * location_mult * penetration_rate * weapon_attenuation
                armor_damage_taken = armor_dmg_value
                new_durability = gear_durability - armor_damage_taken
            else:
                # 公式b: 护甲部分击穿
                ratio = gear_durability / armor_dmg_value
                damage_part1 = base_damage_this * arrow_data["damage_multiplier"] * location_mult * penetration_rate * weapon_attenuation
                damage_part2 = base_damage_this * arrow_data["damage_multiplier"] * location_mult * weapon_attenuation
                damage = ratio * damage_part1 + (1 - ratio) * damage_part2
                armor_damage_taken = gear_durability
                new_durability = Decimal('0.0')
            
            # 更新装备耐久
            new_durability = round_decimal(new_durability, 1)  # 耐久精确到0.1
            if gear_type == "头盔":
                helmet_durability = max(new_durability, Decimal('0.0'))
            else:
                armor_durability = max(new_durability, Decimal('0.0'))
        
        # 四舍五入伤害值
        damage = round_decimal(damage, 2)
        
        # 更新玩家状态
        health -= damage
        total_damage += damage
        total_armor_damage += armor_damage_taken
        
        # 输出结果
        print(f"\n{shot_type}命中 {location}!")
        print(f"造成伤害: {float(damage):.2f}")
        
        if protected:
            armor_damage_taken = round_decimal(armor_damage_taken, 1)
            print(f"{gear_type}损失耐久: {float(armor_damage_taken):.1f}")
            if gear_durability > Decimal('0') and armor_damage_taken >= gear_durability:
                print(f"{gear_type}已击碎!")
        
        if health > Decimal('0'):
            health_rounded = round_decimal(health, 1)
            print(f"剩余生命: {float(health_rounded):.1f}")
            print(f"头盔耐久: {float(helmet_durability):.1f}")
            print(f"护甲耐久: {float(armor_durability):.1f}")
        else:
            print("你死了!")
            print(f"最终头盔耐久: {float(helmet_durability):.1f}")
            print(f"最终护甲耐久: {float(armor_durability):.1f}")
            input("按回车键结束模拟计算")
            break
    
    # 输出统计信息
    print("\n===== 战斗统计 =====")
    print(f"总伤害: {float(total_damage):.2f}")
    print(f"总护甲伤害: {float(total_armor_damage):.1f}")
    print(f"总攻击次数: {total_shots}")
    print(f"总命中次数: {total_hits}")
    print(f"总耗时: {float(total_time_ms/Decimal('1000')):.3f}秒")
    
    print("\n射击方式统计:")
    for shot_type, count in shot_type_count.items():
        if count > 0:
            print(f"{shot_type}: {count}次")
    
    print("\n部位命中统计:")
    for loc, count in location_hits.items():
        if count > 0:
            print(f"{loc}: {count}次")

if __name__ == "__main__":
    main()