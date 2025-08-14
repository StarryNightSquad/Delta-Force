print("本程序由B站繁星攻略组制作")

import sys
from decimal import Decimal, ROUND_HALF_UP
import openpyxl

def get_decimal_input(prompt, min_val, max_val, decimal_places):
    """获取精确的小数输入"""
    while True:
        try:
            value = input(prompt)
            decimal_value = Decimal(value)
            quantized = decimal_value.quantize(
                Decimal('1.' + '0' * decimal_places), 
                rounding=ROUND_HALF_UP
            )
            float_val = float(quantized)
            if min_val <= float_val <= max_val:
                return quantized
            else:
                print(f"输入错误，范围{min_val}到{max_val}，请重新输入。")
        except Exception:
            print("输入错误，请输入有效的数字。")

def get_int_input(prompt, min_val, max_val):
    """获取整数输入"""
    while True:
        try:
            value = int(input(prompt))
            if min_val <= value <= max_val:
                return value
            else:
                print(f"输入错误，范围{min_val}到{max_val}，请重新输入。")
        except ValueError:
            print("输入错误，请输入整数。")

def load_melee_weapons_data():
    """加载近战武器数据"""
    try:
        wb = openpyxl.load_workbook('S5近战武器.xlsx', data_only=True)
        ws = wb['Sheet1']
        
        weapons = []
        
        for row in range(3, ws.max_row + 1):
            weapon_name = ws.cell(row=row, column=1).value
            if not weapon_name or not str(weapon_name).strip():
                continue
                
            # 创建武器数据字典
            weapon_data = {
                'name': str(weapon_name).strip(),
                'damages': [],
                'armor_damages': [],
                'head_multipliers': []
            }
            
            # 基础伤害列：2,3,4 (B,C,D)
            # 护甲伤害列：6,7,8 (F,G,H)
            # 爆头倍率列：10,11,12 (J,K,L)
            
            for i in range(3):  # 三组数据
                # 基础伤害
                damage_col = 2 + i  # B,C,D
                damage = ws.cell(row=row, column=damage_col).value
                
                # 护甲伤害
                armor_col = 6 + i  # F,G,H
                armor_damage = ws.cell(row=row, column=armor_col).value
                
                # 爆头倍率
                head_col = 10 + i  # J,K,L
                head_multiplier = ws.cell(row=row, column=head_col).value
                
                # 处理特殊值
                if damage in (None, "", "/", "//", "N/A"):
                    continue
                    
                try:
                    # 转换为数值
                    weapon_data['damages'].append(float(damage))
                    weapon_data['armor_damages'].append(float(armor_damage))
                    weapon_data['head_multipliers'].append(float(head_multiplier))
                except (ValueError, TypeError):
                    continue
            
            # 只添加有有效数据的武器
            if weapon_data['damages']:
                weapons.append(weapon_data)
                # 调试信息
                print(f"加载武器: {weapon_data['name']}, 连击数: {len(weapon_data['damages'])}")
        
        return weapons
    
    except Exception as e:
        print(f"\n错误: 无法加载近战武器数据 - {e}")
        return []

#读取防具数据
def select_protection(items, item_type):
    """选择防护装备并输入耐久"""
    print(f"\n=== 选择{item_type} ===")
    print("0. 无")
    
    # 按等级分组
    levels = {}
    for item in items:
        level = item['level']
        if level not in levels:
            levels[level] = []
        levels[level].append(item)
    
    # 显示等级选项
    sorted_levels = sorted(levels.keys())
    for i, level in enumerate(sorted_levels, 1):
        print(f"{i}. {level}级{item_type}")
    
    # 获取用户选择，包括0（无）选项
    max_choice = len(sorted_levels)
    level_choice = get_int_input(f"请选择{item_type}等级 (0-{max_choice}): ", 0, max_choice)
    
    # 处理"无"选项
    if level_choice == 0:
        return None, Decimal('0.0')
    
    selected_level = sorted_levels[level_choice - 1]
    
    # 显示该等级下的装备
    level_items = levels[selected_level]
    print(f"\n{selected_level}级{item_type}列表:")
    for i, item in enumerate(level_items, 1):
        print(f"{i}. {item['name']} (最大耐久: {item['max_durability']})")
    
    item_choice = get_int_input(f"请选择{item_type}: ", 1, len(level_items))
    selected_item = level_items[item_choice - 1]
    
    # 输入当前耐久
    max_durability = float(selected_item['max_durability'])
    durability = get_decimal_input(
        f"请输入当前{item_type}耐久 (0.0-{max_durability}): ",
        0.0, max_durability, 1
    )
    
    return selected_item, durability

def load_armor_data():
    """加载护甲和头盔数据"""
    try:
        wb = openpyxl.load_workbook('S5护甲数据.xlsx', data_only=True)
        ws = wb.active
        
        armors = []
        helmets = []
        current_section = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            
            # 检测章节开始
            if cell_value == "护甲":
                current_section = "armor"
                continue
            elif cell_value == "头盔":
                current_section = "helmet"
                continue
                
            if not cell_value or not str(cell_value).strip():
                continue
                
            # 读取数据
            name = str(cell_value).strip()
            level = ws.cell(row=row, column=2).value
            armor_type = ws.cell(row=row, column=3).value
            max_durability = ws.cell(row=row, column=7).value  # G列: 初始上限
            
            if level is None or max_durability is None:
                continue
                
            try:
                level = int(level)
                max_durability = Decimal(str(max_durability))
            except (ValueError, TypeError):
                continue
                
            item_data = {
                'name': name,
                'level': level,
                'max_durability': max_durability
            }
            
            if current_section == "armor":
                # 护甲类型映射: 半甲->1, 全甲->2, 重甲->3
                if armor_type == '半甲':
                    item_data['armor_type'] = 1
                elif armor_type == '全甲':
                    item_data['armor_type'] = 2
                elif armor_type == '重甲':
                    item_data['armor_type'] = 3
                else:
                    continue  # 跳过无效类型
                armors.append(item_data)
                
            elif current_section == "helmet":
                helmets.append(item_data)
                
        return armors, helmets
    
    except Exception as e:
        print(f"\n错误: 无法加载护甲数据 - {e}")
        return [], []

def select_protection(items, item_type):
    """选择防护装备并输入耐久"""
    print(f"\n=== 选择{item_type} ===")
    print("0. 无")
    
    # 按等级分组
    levels = {}
    for item in items:
        level = item['level']
        if level not in levels:
            levels[level] = []
        levels[level].append(item)
    
    # 显示等级选项
    sorted_levels = sorted(levels.keys())
    for i, level in enumerate(sorted_levels, 1):
        print(f"{i}. {level}级{item_type}")
    
    # 获取用户选择，包括0（无）选项
    max_choice = len(sorted_levels)
    level_choice = get_int_input(f"请选择{item_type}等级 (0-{max_choice}): ", 0, max_choice)
    
    # 处理"无"选项
    if level_choice == 0:
        return None, Decimal('0.0')
    
    selected_level = sorted_levels[level_choice - 1]
    
    # 显示该等级下的装备
    level_items = levels[selected_level]
    print(f"\n{selected_level}级{item_type}列表:")
    for i, item in enumerate(level_items, 1):
        print(f"{i}. {item['name']} (最大耐久: {item['max_durability']})")
    
    item_choice = get_int_input(f"请选择{item_type}: ", 1, len(level_items))
    selected_item = level_items[item_choice - 1]
    
    # 输入当前耐久
    max_durability = float(selected_item['max_durability'])
    durability = get_decimal_input(
        f"请输入当前{item_type}耐久 (0.0-{max_durability}): ",
        0.0, max_durability, 1
    )
    
    return selected_item, durability

def main():
    # 初始化参数
    print("=== 近战武器伤害模拟器 ===")
    
    # 加载护甲数据
    print("正在加载护甲数据...")
    armors, helmets = load_armor_data()
    
    # 选择头盔
    helmet_level = 0
    helmet_durability = Decimal('0.0')
    armor_type = 0
    
    if helmets:
        selected_helmet, helmet_durability = select_protection(helmets, "头盔")
        if selected_helmet:
            helmet_level = selected_helmet['level']
            print(f"已选择头盔: {selected_helmet['name']} (等级{helmet_level})")
        else:
            print("已选择：无头盔")
    else:
        print("未找到头盔数据，将使用无头盔设置")
    
    # 选择护甲
    armor_level = 0
    armor_durability = Decimal('0.0')
    armor_type_value = 0
    
    if armors:
        selected_armor, armor_durability = select_protection(armors, "护甲")
        if selected_armor:
            armor_level = selected_armor['level']
            armor_type_value = selected_armor['armor_type']
            print(f"已选择护甲: {selected_armor['name']} (等级{armor_level}, 类型{armor_type_value})")
        else:
            print("已选择：无护甲")
    else:
        print("未找到护甲数据，将使用无护甲设置")
    
    # 设置保护的身体部位
    protected_areas = {
        1: ['胸部', '腹部'],
        2: ['胸部', '腹部', '下腹部'],
        3: ['胸部', '腹部', '下腹部', '大臂']
    }.get(armor_type_value, [])  # 如果没有护甲，返回空列表
     
    # 加载近战武器数据
    print("正在加载近战武器数据...")
    weapons = load_melee_weapons_data()
    if not weapons:
        print("无法加载武器数据，程序退出")
        return
    
    # 选择武器
    print("\n=== 选择近战武器 ===")
    
    # 显示武器列表
    print("\n近战武器列表：")
    for i, weapon in enumerate(weapons, 1):
        hits = len(weapon['damages'])
        print(f"{i}. {weapon['name']} ({hits}连击)")
    
    weapon_choice = get_int_input("输入武器编号：", 1, len(weapons))
    selected_weapon = weapons[weapon_choice - 1]
    max_hits = len(selected_weapon['damages'])
    
    # 初始化状态
    player_health = Decimal('100.0')
    current_helmet_durability = helmet_durability
    current_armor_durability = armor_durability
    protected_areas = {
        1: ['胸部', '腹部'],
        2: ['胸部', '腹部', '下腹部'],
        3: ['胸部', '腹部', '下腹部', '大臂']
    }[armor_type]
    limb_areas = ['大臂', '小臂', '大腿', '小腿']
    valid_parts = ['头部', '胸部', '腹部', '大臂', '小臂', '大腿', '小腿', '下腹部', '未命中']
    
    hit_count = 0
    sequence_index = 0  # 当前连击序号
    total_damage = Decimal('0.0')
    total_armor_damage = Decimal('0.0')
    hit_statistics = {part: 0 for part in valid_parts}
    
    print("\n=== 开始模拟计算 ===")
    print(f"武器: {selected_weapon['name']}, 最大连击: {max_hits}")
    while True:
        # 输入命中部位
        while True:
            hit_part = input(f"\n第{sequence_index+1}击 - 输入命中部位（头部/胸部/腹部/下腹部/大臂/小臂/大腿/小腿/未命中）：").strip()
            if hit_part.lower() == 'exit':
                return
            if hit_part in valid_parts:
                hit_statistics[hit_part] += 1
                hit_count += 1
                break
            print("无效输入，请重新输入。")
        
        # 处理未命中
        if hit_part == '未命中':
            print("\n=== 计算结果 ===")
            print("本次攻击未命中")
            # 移动到下一击
            sequence_index = (sequence_index + 1) % max_hits
            continue
        
        # 获取当前连击数据
        hit_index = sequence_index % max_hits
        base_damage = Decimal(str(selected_weapon['damages'][hit_index]))
        armor_damage_value = Decimal(str(selected_weapon['armor_damages'][hit_index]))
        head_multiplier = Decimal(str(selected_weapon['head_multipliers'][hit_index]))
        
        # 计算部位倍率
        if hit_part == '头部':
            part_multiplier = head_multiplier
        elif hit_part in limb_areas:
            part_multiplier = Decimal('0.8')
        else:
            part_multiplier = Decimal('1.0')
        
        # 伤害计算逻辑
        is_protected = False
        protector_level = 0
        protector_type = None
        current_protector_durability = Decimal('0.0')
        
        # 判断保护状态
        if hit_part == '头部':
            if helmet_level > 0 and current_helmet_durability > Decimal('0'):
                is_protected = True
                protector_level = helmet_level
                protector_type = 'helmet'
                current_protector_durability = current_helmet_durability
        else:
            if armor_level > 0 and current_armor_durability > Decimal('0') and hit_part in protected_areas:
                is_protected = True
                protector_level = armor_level
                protector_type = 'armor'
                current_protector_durability = current_armor_durability
        
        # 计算实际伤害
        final_damage = Decimal('0.0')
        armor_damage_dealt = Decimal('0.0')
        protector_destroyed = False
        
        if is_protected:
            # 护甲减伤计算
            if protector_level >= 2:
                damage_reduction = Decimal('1.0') - (Decimal(str(protector_level)) * Decimal('0.05'))
                base_damage_after_reduction = base_damage * damage_reduction
            else:
                base_damage_after_reduction = base_damage
            
            # 计算护甲伤害
            if current_protector_durability >= armor_damage_value:
                # 护甲足够承受全部伤害
                armor_damage_dealt = armor_damage_value
                final_damage = base_damage_after_reduction * part_multiplier
            else:
                # 护甲不足以承受全部伤害 - 使用部分减伤计算
                ratio = current_protector_durability / armor_damage_value
                armor_damage_dealt = current_protector_durability
                
                # 计算两部分伤害
                protected_damage = ratio * base_damage_after_reduction * part_multiplier
                unprotected_damage = (1 - ratio) * base_damage * part_multiplier
                final_damage = protected_damage + unprotected_damage
                
                protector_destroyed = True
            
            # 更新耐久
            if protector_type == 'helmet':
                current_helmet_durability -= armor_damage_dealt
                if current_helmet_durability <= Decimal('0'):
                    current_helmet_durability = Decimal('0.0')
            else:
                current_armor_durability -= armor_damage_dealt
                if current_armor_durability <= Decimal('0'):
                    current_armor_durability = Decimal('0.0')
        else:
            # 未受保护
            final_damage = base_damage * part_multiplier
        
        # 更新统计
        total_damage += final_damage
        total_armor_damage += armor_damage_dealt
        player_health -= final_damage
        
        # 四舍五入伤害值
        final_damage = final_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        armor_damage_dealt = armor_damage_dealt.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
        
        # 输出结果
        print("\n=== 计算结果 ===")
        print(f"连击: 第{hit_index+1}击")
        print(f"基础伤害: {base_damage}, 部位倍率: {part_multiplier}")
        
        if is_protected:
            print(f"护甲减伤: 等级{protector_level} {'(有减伤)' if protector_level >= 2 else '(无减伤)'}")
            
            if protector_destroyed:
                print(f"{'头盔' if protector_type == 'helmet' else '护甲'}被击碎!")
            
            if current_protector_durability < armor_damage_value:
                ratio_val = (current_protector_durability / armor_damage_value).quantize(Decimal('0.01'))
                print(f"护甲部分吸收: 吸收率{ratio_val} (耐久不足)")
            
            print(f"{'头盔' if protector_type == 'helmet' else '护甲'}损失耐久: {armor_damage_dealt}")
        
        print(f"造成伤害: {final_damage}")
        print(f"剩余生命值: {player_health.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}")
        print(f"剩余头盔耐久: {current_helmet_durability}")
        print(f"剩余护甲耐久: {current_armor_durability}")
        
        # 死亡处理
        if player_health <= Decimal('0'):
            print("\n=== 最终统计 ===")
            print(f"总造成伤害: {total_damage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}")
            print(f"总护甲伤害: {total_armor_damage.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}")
            
            print("\n命中统计:")
            total_shots = sum(hit_statistics.values())
            valid_hits = total_shots - hit_statistics['未命中']
            display_order = ['未命中', '头部', '胸部', '腹部', '下腹部', '大臂', '小臂', '大腿', '小腿']
            for part in display_order:
                count = hit_statistics.get(part, 0)
                if count > 0:
                    print(f"{part.ljust(5)}: {count}次")
            
            print(f"\n有效命中次数: {valid_hits}次")
            print(f"总攻击次数: {total_shots}次")
            
            input("\n按回车键结束模拟计算...")
            break
        
        # 移动到下一击
        sequence_index = (sequence_index + 1) % max_hits

if __name__ == "__main__":
    main()
