print("本程序由B站繁星攻略组制作，感谢B站用户Dec128与乂丶z提供的数据支持")

import math
import re
import os
from decimal import Decimal, getcontext, ROUND_FLOOR, ROUND_HALF_UP
from openpyxl import load_workbook

# 设置decimal的精度环境
getcontext().prec = 10
getcontext().rounding = ROUND_HALF_UP

def validate_float_input(prompt, min_val, max_val, decimal_places, value_name, reference_value=None, reference_type=None):
    """验证浮点数输入，确保格式、范围和精度"""
    while True:
        value = input(prompt)
        try:
            # 格式校验（最多指定小数位）
            pattern = rf'^\d+(\.\d{{0,{decimal_places}}})?$'
            if not re.match(pattern, value):
                raise ValueError(f"最多允许{decimal_places}位小数")
            
            num = float(value)
            if num < min_val:
                raise ValueError(f"不能小于{min_val}")
            if num > max_val:
                raise ValueError(f"不能大于{max_val}")
            
            # 额外的验证规则
            if reference_value is not None:
                if reference_type == "initial_upper" and num > reference_value:
                    raise ValueError(f"不能高于初始上限({reference_value})")
                elif reference_type == "current_upper" and num > reference_value:
                    raise ValueError(f"不能高于当前上限({reference_value})")
            
            # 规范小数位数（不四舍五入，仅截断）
            num = float(f"{num:.{decimal_places}f")
            return num
        except ValueError as e:
            print(f"{value_name}输入错误：{e}，请重新输入")

def validate_int_input(prompt, min_val, max_val, value_name):
    """验证整数输入，确保范围和格式"""
    while True:
        value = input(prompt)
        try:
            num = int(value)
            if num < min_val:
                raise ValueError(f"不能小于{min_val}")
            if num > max_val:
                raise ValueError(f"不能大于{max_val}")
            return num
        except ValueError:
            print(f"{value_name}输入错误：必须为整数，请重新输入")

def load_equipment_data(file_path):
    """从Excel文件加载装备数据"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
        
        armor_data = {3: [], 4: [], 5: [], 6: []}  # 护甲数据按等级分组
        helmet_data = {3: [], 4: [], 5: [], 6: []}  # 头盔数据按等级分组
        
        # 从第4行开始读取数据（跳过前3行标题）
        for row in range(4, sheet.max_row + 1):
            # 读取装备名称
            name = sheet.cell(row=row, column=1).value
            if not name or name == " " or name == "":
                continue
                
            # 读取防护等级
            level = sheet.cell(row=row, column=2).value
            if level is None or level not in [3, 4, 5, 6]:
                continue
                
            # 读取装备类型
            armor_type = sheet.cell(row=row, column=3).value
            if not armor_type:
                continue
                
            # 读取初始上限
            initial_upper = sheet.cell(row=row, column=7).value
            if initial_upper is None:
                continue
                
            # 读取维修损耗
            repair_loss = sheet.cell(row=row, column=9).value
            if repair_loss is None:
                continue
                
            # 读取维修单价
            repair_price = sheet.cell(row=row, column=10).value
            if repair_price is None:
                continue
                
            # 读取维修效率数据（K、M、O、Q列）
            repair_efficiency = []
            for col in [11, 13, 15, 17]:
                value = sheet.cell(row=row, column=col).value
                if value and isinstance(value, (int, float)):
                    repair_efficiency.append(value)
            
            # 创建装备数据字典
            equipment = {
                'name': name.strip(),
                'level': level,
                'type': armor_type.strip(),
                'initial_upper': Decimal(str(initial_upper)),  # 使用Decimal
                'repair_loss': Decimal(str(repair_loss)),      # 使用Decimal
                'repair_price': Decimal(str(repair_price)),    # 使用Decimal
                'repair_efficiency': repair_efficiency
            }
            
            # 根据类型分类存储
            if armor_type in ["半甲", "全甲", "重甲"]:
                armor_data[level].append(equipment)
            elif armor_type in ["无", "有"]:
                helmet_data[level].append(equipment)
        
        return armor_data, helmet_data
    
    except Exception as e:
        raise RuntimeError(f"读取Excel文件失败: {e}")

def main():
    # 加载装备数据
    try:
        file_path = "S5护甲数据.xlsx"
        armor_data, helmet_data = load_equipment_data(file_path)
        print(f"成功加载装备数据，共找到 {sum(len(v) for v in armor_data.values())} 件护甲和 {sum(len(v) for v in helmet_data.values())} 件头盔")
    except Exception as e:
        print(f"加载装备数据失败: {e}")
        input("\n按回车结束程序")
        return
    
    print("\n=== 装备维修计算器 ===")
    
    # 选择装备类型
    print("\n请选择装备类型:")
    print("1. 护甲")
    print("2. 头盔")
    item_choice = validate_int_input("请输入选择(1-2): ", 1, 2, "装备类型")
    
    item_type = "护甲" if item_choice == 1 else "头盔"
    data_source = armor_data if item_choice == 1 else helmet_data
    
    # 选择装备等级
    print("\n请选择装备等级:")
    print("3. 3级装备")
    print("4. 4级装备")
    print("5. 5级装备")
    print("6. 6级装备")
    level_choice = validate_int_input("请输入选择(3-6): ", 3, 6, "装备等级")
    
    # 获取该等级下的装备列表
    equipment_list = data_source.get(level_choice, [])
    if not equipment_list:
        print(f"\n没有找到{level_choice}级{item_type}数据")
        input("\n按回车结束程序")
        return
    
    # 显示装备列表供选择
    print(f"\n请选择{level_choice}级{item_type}:")
    for idx, eq in enumerate(equipment_list, 1):
        print(f"{idx}. {eq['name']} (初始上限: {eq['initial_upper']}, 维修损耗: {eq['repair_loss']})")
    
    # 选择具体装备
    eq_choice = validate_int_input(f"请输入选择(1-{len(equipment_list)}): ", 1, len(equipment_list), "装备")
    selected_eq = equipment_list[eq_choice - 1]
    
    # 显示装备详细信息
    print("\n=== 装备详细信息 ===")
    print(f"装备名称: {selected_eq['name']}")
    print(f"装备类型: {selected_eq['type']}")
    print(f"初始上限: {selected_eq['initial_upper']}")
    print(f"维修损耗: {selected_eq['repair_loss']}")
    print(f"维修单价: {selected_eq['repair_price']}")
    
    # 显示维修效率信息
    if selected_eq['repair_efficiency']:
        print("\n维修效率信息:")
        for idx, eff in enumerate(selected_eq['repair_efficiency'], 1):
            print(f"维修效率{idx}: {eff}")
    
    # 输入当前上限和剩余耐久（带额外验证）
    print("\n请提供装备当前状态:")
    current_upper = validate_float_input(
        f"请输入当前上限（1-{selected_eq['initial_upper']}，最多1位小数）: ",
        1, 150, 1, "当前上限",
        reference_value=float(selected_eq['initial_upper']),
        reference_type="initial_upper"
    )
    
    remaining_durability = validate_float_input(
        f"请输入剩余耐久（0-{current_upper}，最多1位小数）: ",
        0, 150, 1, "剩余耐久",
        reference_value=current_upper,
        reference_type="current_upper"
    )
    
    # 提取装备属性并转换为Decimal
    initial_upper = selected_eq['initial_upper']
    repair_loss = selected_eq['repair_loss']
    repair_price = selected_eq['repair_price']
    
    # 处理当前上限（去尾法取整）并转换为Decimal
    current_upper_processed = Decimal(str(int(current_upper)))
    remaining_durability_dec = Decimal(str(remaining_durability))

    # 维修可行性判定
    print("\n=== 维修可行性判定 ===")
    if item_type == "护甲" and current_upper_processed < 10:
        print(f"当前护甲上限({current_upper_processed})小于10，不可维修")
    elif item_type == "头盔" and current_upper_processed < 5:
        print(f"当前头盔上限({current_upper_processed})小于5，不可维修")
    else:
        print("装备符合维修条件，开始计算...")
        
        try:
            # 计算公式组件 - 使用Decimal进行精确计算
            term1 = (current_upper_processed - remaining_durability_dec) / current_upper_processed
            
            # 对数计算安全校验
            if current_upper_processed <= 0 or initial_upper <= 0:
                raise ValueError("对数计算参数必须大于0")
                
            # 计算对数部分（使用浮点数）
            log_value = float(current_upper_processed) / float(initial_upper)
            if log_value <= 0:
                raise ValueError("对数计算参数必须大于0")
                
            log_term = math.log10(log_value)
            
            # 使用Decimal进行后续计算
            term2 = repair_loss - Decimal(str(log_term))
            
            # 计算维修后上限
            repaired_upper = current_upper_processed - current_upper_processed * term1 * term2 * 1.25
            
            # 确保结果有效
            if repaired_upper.is_nan() or repaired_upper.is_infinite():
                raise ValueError("计算结果无效")

        except Exception as e:
            print(f"\n计算错误：{e}")
            input("\n按回车结束程序")
            return

        # 结果处理 - 使用Decimal的整数转换
        # 去尾法取整（向下取整）
        final_upper = repaired_upper.to_integral_value(rounding=ROUND_FLOOR)
        if final_upper < 1:
            final_upper = Decimal('1')
        
        # 计算维修花费（剩余耐久取整）
        remaining_int = Decimal(str(int(remaining_durability)))  # 去尾取整
        repair_cost = (final_upper - remaining_int + Decimal('1')) * repair_price * 1.25
        
        # 花费不能为负
        if repair_cost < 0:
            repair_cost = Decimal('0')
        else:
            # 货币值取整（四舍五入到整数）
            repair_cost = repair_cost.to_integral_value(rounding=ROUND_HALF_UP)

        print("\n=== 计算结果 ===")
        print(f"装备名称: {selected_eq['name']}")
        print(f"装备等级: {level_choice}级")
        print(f"初始上限: {initial_upper}")
        print(f"当前上限(输入): {current_upper} → 处理值: {current_upper_processed}")
        print(f"剩余耐久: {remaining_durability_dec}")
        print(f"维修损耗: {repair_loss}")
        print(f"维修后耐久上限: {final_upper}")
        print(f"维修花费: {repair_cost:,}")  # 千位分隔符格式化

    input("\n按回车结束程序")

if __name__ == "__main__":
    main()
