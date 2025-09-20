print("本程序由B站繁星攻略组制作")

import math
import openpyxl
import os
import sys
from decimal import Decimal, ROUND_HALF_UP

def load_weapon_data(file_path):
    """从Excel文件加载武器数据，跳过霰弹枪"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb["战场模式"]
        
        weapons = []
        current_category = ""
        row_idx = 3  # 从第3行开始读取数据
        
        while row_idx <= 100:  # 安全限制
            # 读取行数据
            category_cell = sheet.cell(row=row_idx, column=1).value
            name_cell = sheet.cell(row=row_idx, column=2).value
            
            # 更新当前武器类别
            if category_cell:
                current_category = category_cell
            
            # 跳过霰弹枪和空行
            if current_category == "霰弹枪" or not name_cell:
                row_idx += 1
                continue
            
            # 收集武器数据
            weapon = {
                "category": current_category,
                "name": name_cell,
                "fire_mode": sheet.cell(row=row_idx, column=4).value or 1,
                "trigger_delay": sheet.cell(row=row_idx, column=5).value or 0,
                "rpm": sheet.cell(row=row_idx, column=6).value or 0,
                "shooting_interval": sheet.cell(row=row_idx, column=7).value or 0,
                # 基础伤害在第9列(I列)
                "base_damage": sheet.cell(row=row_idx, column=9).value or 0,
                # 部位倍率在第12-18列(L-R列)
                "head_mult": sheet.cell(row=row_idx, column=12).value or 0,  # L列
                "chest_mult": sheet.cell(row=row_idx, column=13).value or 0,  # M列
                "abdomen_mult": sheet.cell(row=row_idx, column=14).value or 0,  # N列
                "upper_arm_mult": sheet.cell(row=row_idx, column=15).value or 0,  # O列
                "forearm_mult": sheet.cell(row=row_idx, column=16).value or 0,  # P列
                "thigh_mult": sheet.cell(row=row_idx, column=17).value or 0,  # Q列
                "calf_mult": sheet.cell(row=row_idx, column=18).value or 0,  # R列
                "decay_distances": [],
                "decay_multipliers": []
            }
            
            # 处理射击间隔 - 如果为空则计算
            if not weapon["shooting_interval"] and weapon["rpm"]:
                weapon["shooting_interval"] = round(60000 / weapon["rpm"], 2)
            
            # 收集衰减数据（最多4组）
            for i in range(4):
                dist_col = 19 + i * 2  # S, U, W, Y列
                mult_col = 20 + i * 2  # T, V, X, Z列
                
                dist = sheet.cell(row=row_idx, column=dist_col).value
                mult = sheet.cell(row=row_idx, column=mult_col).value
                
                # 跳过无效数据
                if dist in ["/", None] or mult in ["/", None]:
                    continue
                
                # 确保是数值类型
                try:
                    dist_val = float(dist)
                    mult_val = float(mult)
                    weapon["decay_distances"].append(dist_val)
                    weapon["decay_multipliers"].append(mult_val)
                except (ValueError, TypeError):
                    continue
            
            weapons.append(weapon)
            row_idx += 1
        
        return weapons
    
    except Exception as e:
        print(f"读取武器数据时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return []

def get_decay_multiplier(weapon, distance):
    """根据距离获取衰减倍率"""
    if not weapon["decay_distances"]:
        return 1.0
    
    # 确保衰减距离有序
    sorted_distances = sorted(weapon["decay_distances"])
    sorted_multipliers = [x for _, x in sorted(zip(weapon["decay_distances"], weapon["decay_multipliers"]))]
    
    # 在第一个衰减距离前，无衰减
    if distance <= sorted_distances[0]:
        return 1.0
    
    # 检查距离落在哪个衰减区间
    for i, dist in enumerate(sorted_distances):
        if distance <= dist:
            return sorted_multipliers[i-1] if i > 0 else 1.0
    
    # 距离超过最后一个衰减距离，使用最后一个衰减倍率
    return sorted_multipliers[-1]

def main():
    print("战场武器伤害计算器")
    print("=" * 50)
    
    # 检查文件是否存在
    file_path = "S6战场武器.xlsx"
    if not os.path.exists(file_path):
        print(f"错误: 武器数据文件 '{file_path}' 不存在")
        print("请将Excel文件放在程序同一目录下")
        input("按回车键退出...")
        return
    
    # 加载武器数据
    print("正在加载武器数据...")
    weapons = load_weapon_data(file_path)
    if not weapons:
        print("未找到有效的武器数据")
        input("按回车键退出...")
        return
    
    # 按武器类型分组
    weapon_categories = {}
    for weapon in weapons:
        category = weapon["category"]
        if category not in weapon_categories:
            weapon_categories[category] = []
        weapon_categories[category].append(weapon)
    
    # 显示武器类型供选择
    print("\n可用武器类型:")
    print("（复合弓 点射武器与霰弹枪不适用本程序）")
    sorted_categories = sorted(weapon_categories.keys())
    for idx, category in enumerate(sorted_categories, 1):
        print(f"{idx}. {category}")
    
    # 选择武器类型
    try:
        cat_choice = int(input("\n请选择武器类型 (输入编号): "))
        if cat_choice < 1 or cat_choice > len(sorted_categories):
            print("无效的选择!")
            return
        
        selected_category = sorted_categories[cat_choice - 1]
        category_weapons = weapon_categories[selected_category]
        
        print(f"\n已选择: {selected_category}")
    except ValueError:
        print("请输入有效的数字!")
        return
    
    # 显示具体武器供选择
    print(f"\n{selected_category}可用武器:")
    for idx, weapon in enumerate(category_weapons, 1):
        print(f"{idx}. {weapon['name']}")
    
    # 选择具体武器
    try:
        weapon_choice = int(input("\n请选择武器 (输入编号): "))
        if weapon_choice < 1 or weapon_choice > len(category_weapons):
            print("无效的选择!")
            return
        
        selected_weapon = category_weapons[weapon_choice - 1]
        print(f"\n已选择: {selected_weapon['name']}")
    except ValueError:
        print("请输入有效的数字!")
        return
    
    # 输入目标距离
    try:
        distance = float(input("\n请输入目标距离 (0-400米): "))
        if distance < 0 or distance > 400:
            print("距离必须在0-400米之间!")
            return
    except ValueError:
        print("请输入有效的数字!")
        return
    
    # 计算衰减倍率
    decay_multiplier = get_decay_multiplier(selected_weapon, distance)
    
    # 获取武器数据
    fire_mode = selected_weapon["fire_mode"]
    trigger_delay = selected_weapon["trigger_delay"]
    rpm = selected_weapon["rpm"]
    shooting_interval = selected_weapon["shooting_interval"]
    base_damage = selected_weapon["base_damage"]
    
    # 部位倍率
    multipliers = {
        "头部": selected_weapon["head_mult"],
        "胸部": selected_weapon["chest_mult"],
        "腹部": selected_weapon["abdomen_mult"],
        "大臂": selected_weapon["upper_arm_mult"],
        "小臂": selected_weapon["forearm_mult"],
        "大腿": selected_weapon["thigh_mult"],
        "小腿": selected_weapon["calf_mult"]
    }
    
    # 计算并显示结果
    print("\n" + "=" * 50)
    print(f"武器: {selected_weapon['name']} ({selected_weapon['category']})")
    print(f"目标距离: {distance}米, 衰减倍率: {decay_multiplier:.3f}")
    print(f"射击模式: {'全自动' if fire_mode == 1 else '半自动'}")
    print(f"基础伤害: {base_damage}, 射速: {rpm}RPM, 射击间隔: {shooting_interval:.2f}ms")
    print(f"扳机延迟: {trigger_delay}ms")
    
    # 打印部位倍率用于验证
    print("\n部位倍率:")
    print(f"头部: {multipliers['头部']} | 胸部: {multipliers['胸部']} | 腹部: {multipliers['腹部']}")
    print(f"大臂: {multipliers['大臂']} | 小臂: {multipliers['小臂']} | 大腿: {multipliers['大腿']} | 小腿: {multipliers['小腿']}")
    
    print("\n各部位致死次数及耗时:")
    print("-" * 50)
    body_parts = ["头部", "胸部", "腹部", "大臂", "小臂", "大腿", "小腿"]
    
    for part in body_parts:
        multiplier = multipliers[part]
        dmg_per_shot = base_damage * multiplier * decay_multiplier
        
        if dmg_per_shot <= 0:
            hits_str = "无法致死"
            time_str = "无法计算"
        else:
            hits = math.ceil(100 / dmg_per_shot)
            
            # 根据射击模式计算耗时
            if fire_mode == 1:  # 全自动
                time_ms = trigger_delay + shooting_interval * (hits - 1)
            else:  # 半自动
                time_ms = trigger_delay * hits + shooting_interval * (hits - 1)
            
            time_ms = round(time_ms, 2)
            hits_str = f"{hits}次"
            time_str = f"{time_ms:.2f}ms"
        
        print(f"{part.ljust(4)}：{hits_str.ljust(8)}，耗时{time_str}")
    
    print("=" * 50)
    input("\n按回车键结束模拟计算...")

if __name__ == "__main__":
    main()