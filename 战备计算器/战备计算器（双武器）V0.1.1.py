import openpyxl
import time
import heapq
from collections import defaultdict

def load_equipment_data(file_path):
    """加载装备数据（护甲、头盔、胸挂、背包），并过滤掉市场价值≥战备价值+2000的物品"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb["护甲数据"]
    
    categories = {
        "护甲": [],
        "头盔": [],
        "胸挂": [],
        "背包": []
    }
    
    current_category = None
    overpriced_items = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # 跳过空行
            continue
            
        if row[0] in categories:  # 检测到类别标题
            current_category = row[0]
            continue
            
        if current_category and row[1] and (row[2] is not None or row[3] is not None):
            # 处理有价格数据的行
            name = row[1]
            market_price = row[2] or 0
            readiness_value = row[3] or 0
            
            # 检查是否市场价值≥战备价值+2000
            if market_price >= readiness_value + 2000:
                overpriced_items.append({
                    "category": current_category,
                    "name": name,
                    "market_price": market_price,
                    "readiness_value": readiness_value
                })
                continue  # 跳过这个物品
            
            categories[current_category].append({
                "name": name,
                "market_price": market_price,
                "readiness_value": readiness_value
            })
    
    return categories, overpriced_items

def load_weapon_data(file_path):
    """加载武器数据（非手枪武器和手枪武器），并过滤掉市场价值≥战备价值+2000的物品"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb["夺金模式"]
    
    weapons = defaultdict(list)
    current_category = None
    overpriced_weapons = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # 跳过空行
            continue
            
        if row[0] and row[0] != "手枪":  # 检测到武器类别（排除手枪）
            current_category = row[0]
            continue
            
        if current_category and row[1] and (row[2] is not None or row[3] is not None):
            # 处理有价格数据的行
            name = row[1]
            market_price = row[2] or 0
            readiness_value = row[3] or 0
            
            # 检查是否市场价值≥战备价值+2000
            if market_price >= readiness_value + 2000:
                overpriced_weapons.append({
                    "category": current_category,
                    "name": name,
                    "market_price": market_price,
                    "readiness_value": readiness_value
                })
                continue  # 跳过这个武器
            
            weapons[current_category].append({
                "name": name,
                "market_price": market_price,
                "readiness_value": readiness_value
            })
    
    # 单独处理手枪类别
    pistol_rows = [
        (7, "M1911"),
        (8, "G17"),
        (9, "G18"),
        (10, "沙漠之鹰"),
        (11, ".357左轮"),
        (12, "QSZ92G"),
        (13, "R93")
    ]
    
    for row_idx, name in pistol_rows:
        row = sheet[row_idx]
        market_price = row[2].value if row[2] and row[2].value is not None else 0
        readiness_value = row[3].value if row[3] and row[3].value is not None else 0
        
        # 检查是否市场价值≥战备价值+2000
        if market_price >= readiness_value + 2000:
            overpriced_weapons.append({
                "category": "手枪",
                "name": name,
                "market_price": market_price,
                "readiness_value": readiness_value
            })
            continue  # 跳过这个武器
        
        weapons["手枪"].append({
            "name": name,
            "market_price": market_price,
            "readiness_value": readiness_value
        })
    
    return weapons, overpriced_weapons

def filter_items_by_target_value(items, target_value, item_type):
    """过滤掉市场价值≥目标战备值的物品"""
    filtered_items = []
    removed_items = []
    
    for item in items:
        if item["market_price"] >= target_value:
            removed_items.append({
                "type": item_type,
                "name": item["name"],
                "market_price": item["market_price"],
                "readiness_value": item["readiness_value"]
            })
        else:
            filtered_items.append(item)
    
    return filtered_items, removed_items

def generate_combinations(target_value, equipment_data, weapon_data):
    """生成所有装备组合"""
    start_time = time.time()
    
    # 准备槽位数据（包括空选项）
    none_option = {"name": "无", "market_price": 0, "readiness_value": 0}
    
    slots = {
        "weapon1": [none_option],
        "weapon2": [none_option],
        "pistol": [none_option],
        "helmet": [none_option],
        "armor": [none_option],
        "chest": [none_option],
        "backpack": [none_option]
    }
    
    # 填充武器槽（非手枪武器）
    for category in weapon_data:
        if category != "手枪":
            for weapon in weapon_data[category]:
                slots["weapon1"].append(weapon)
                slots["weapon2"].append(weapon)
    
    # 填充手枪槽
    if "手枪" in weapon_data:
        slots["pistol"].extend(weapon_data["手枪"])
    
    # 填充装备槽
    slots["helmet"].extend(equipment_data["头盔"])
    slots["armor"].extend(equipment_data["护甲"])
    slots["chest"].extend(equipment_data["胸挂"])
    slots["backpack"].extend(equipment_data["背包"])
    
    # 生成所有组合
    valid_combinations = []
    counter = 0  # 用于生成唯一标识符
    
    for w1 in slots["weapon1"]:
        for w2 in slots["weapon2"]:
            # 限制：如果武器槽1是"无"，那么武器槽2也必须是"无"
            if w1["name"] == "无" and w2["name"] != "无":
                continue
                
            for pistol in slots["pistol"]:
                for helmet in slots["helmet"]:
                    for armor in slots["armor"]:
                        for chest in slots["chest"]:
                            for backpack in slots["backpack"]:
                                total_value = (
                                    w1["readiness_value"] +
                                    w2["readiness_value"] +
                                    pistol["readiness_value"] +
                                    helmet["readiness_value"] +
                                    armor["readiness_value"] +
                                    chest["readiness_value"] +
                                    backpack["readiness_value"]
                                )
                                
                                if total_value >= target_value:
                                    total_cost = (
                                        w1["market_price"] +
                                        w2["market_price"] +
                                        pistol["market_price"] +
                                        helmet["market_price"] +
                                        armor["market_price"] +
                                        chest["market_price"] +
                                        backpack["market_price"]
                                    )
                                    
                                    counter += 1
                                    combination = {
                                        "cost": total_cost,
                                        "value": total_value,
                                        "items": {
                                            "武器槽1": w1,
                                            "武器槽2": w2,
                                            "手枪槽": pistol,
                                            "头盔槽": helmet,
                                            "护甲槽": armor,
                                            "弹挂槽": chest,
                                            "背包槽": backpack
                                        },
                                        "id": counter  # 添加唯一标识符
                                    }
                                    
                                    # 使用堆维护最小成本的3个组合
                                    # 添加唯一标识符避免字典比较
                                    if len(valid_combinations) < 3:
                                        heapq.heappush(valid_combinations, (-combination["cost"], combination["id"], combination))
                                    else:
                                        heapq.heappushpop(valid_combinations, (-combination["cost"], combination["id"], combination))
    
    # 按成本升序排序
    valid_combinations = sorted([c[2] for c in valid_combinations], key=lambda x: x["cost"])
    
    elapsed_time = time.time() - start_time
    return valid_combinations, elapsed_time

def main():
    # 加载数据
    equipment_data, overpriced_equipment = load_equipment_data("S5装备价格.xlsx")
    weapon_data, overpriced_weapons = load_weapon_data("S5武器价格.xlsx")
    
    # 显示被排除的物品（市场价值≥战备价值+2000）
    if overpriced_equipment or overpriced_weapons:
        print("以下物品因市场价值≥战备价值+2000而被排除:")
        
        for item in overpriced_equipment:
            print(f"  [{item['category']}] {item['name']}: 市场价值={item['market_price']}, 战备价值={item['readiness_value']}")
        
        for item in overpriced_weapons:
            print(f"  [{item['category']}] {item['name']}: 市场价值={item['market_price']}, 战备价值={item['readiness_value']}")
        
        print("\n" + "-" * 50 + "\n")
    
    # 获取用户输入
    target_value = int(input("请输入目标战备价值: "))
    
    # 新限制：过滤掉市场价值≥目标战备值的物品
    removed_by_target = []
    
    # 过滤装备数据
    for category in equipment_data:
        filtered, removed = filter_items_by_target_value(equipment_data[category], target_value, category)
        equipment_data[category] = filtered
        removed_by_target.extend(removed)
    
    # 过滤武器数据
    for category in weapon_data:
        filtered, removed = filter_items_by_target_value(weapon_data[category], target_value, category)
        weapon_data[category] = filtered
        removed_by_target.extend(removed)
    
    # 显示因新限制被排除的物品
    if removed_by_target:
        print(f"\n以下物品因市场价值≥目标战备值({target_value})而被排除:")
        for item in removed_by_target:
            print(f"  [{item['type']}] {item['name']}: 市场价值={item['market_price']}, 战备价值={item['readiness_value']}")
        print("\n" + "-" * 50 + "\n")
    
    # 计算组合
    combinations, elapsed_time = generate_combinations(target_value, equipment_data, weapon_data)
    
    # 输出结果
    print(f"\n计算完成，用时: {elapsed_time:.2f}秒")
    
    if not combinations:
        print("未找到满足条件的组合")
    else:
        print(f"找到 {len(combinations)} 个满足条件的组合，市场价值最低的3个方案:")
        
        for i, combo in enumerate(combinations[:3], 1):
            print(f"\n方案 #{i} - 总市场价值: {combo['cost']}, 总战备价值: {combo['value']}")
            
            for slot, item in combo["items"].items():
                # 输出每个物品的市场价值和战备价值
                print(f"  {slot}: {item['name']}（市场价值={item['market_price']}, 战备价值={item['readiness_value']}）")
    
    input("\n按Enter键退出程序...")

if __name__ == "__main__":
    main()