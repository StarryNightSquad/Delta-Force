print("本程序由B站繁星攻略组制作")
print("卡战备计算程序数据量较大，加载慢与计算慢很正常。耐心等待")

import openpyxl
import time
import heapq
from collections import defaultdict

def load_equipment_data(file_path, filter_overpriced=True, filter_zero_price=True):
    """加载装备数据（护甲、头盔、胸挂、背包）
    filter_overpriced: 是否过滤高价物品（市场价值≥战备价值+2000）
    filter_zero_price: 是否过滤市场价值为0的物品
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb["护甲数据"]
    
    categories = {
        "护甲": [],
        "头盔": [],
        "胸挂": [],
        "背包": []
    }
    
    current_category = None
    overpriced_items = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # 跳过空行
            continue
            
        if row[0] in categories:  # 检测到类别标题
            current_category = row[0]
            continue
            
        if current_category and row[1] and (row[2] is not None or row[3] is not None):
            name = row[1]
            market_price = row[2] or 0
            readiness_value = row[3] or 0
            quality = row[6] or 0  # 获取装备品质
            
            # 跳过市场价值为0的装备（如果启用过滤）
            if filter_zero_price and market_price == 0:
                continue
                
            # 根据参数决定是否过滤高价物品
            if filter_overpriced and market_price >= readiness_value + 2000:
                overpriced_items.append({
                    "category": current_category,
                    "name": name,
                    "market_price": market_price,
                    "readiness_value": readiness_value,
                    "quality": quality
                })
                continue
                
            categories[current_category].append({
                "name": name,
                "market_price": market_price,
                "readiness_value": readiness_value,
                "quality": quality  # 存储品质信息
            })
    
    return categories, overpriced_items

def load_weapon_data(file_path, filter_overpriced=True, filter_zero_price=True):
    """加载武器数据（非手枪武器和手枪武器）
    filter_overpriced: 是否过滤高价物品（市场价值≥战备价值+2000）
    filter_zero_price: 是否过滤市场价值为0的物品
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb["夺金模式"]
    
    weapons = defaultdict(list)
    current_category = None
    overpriced_weapons = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # 跳过空行
            continue
            
        if row[0] and row[0] != "手枪":  # 检测到武器类别（排除手枪）
            current_category = row[0]
            continue
            
        if current_category and row[1] and (row[2] is not None or row[3] is not None):
            name = row[1]
            market_price = row[2] or 0
            readiness_value = row[3] or 0
            
            # 跳过市场价值为0的武器（如果启用过滤）
            if filter_zero_price and market_price == 0:
                continue
                
            if filter_overpriced and market_price >= readiness_value + 2000:
                overpriced_weapons.append({
                    "category": current_category,
                    "name": name,
                    "market_price": market_price,
                    "readiness_value": readiness_value
                })
                continue
                
            weapons[current_category].append({
                "name": name,
                "market_price": market_price,
                "readiness_value": readiness_value
            })
    
    # 单独处理手枪类别
    pistol_rows = [
        (7, "M1911"),
        (8, "G17"),
        (9, "G18"),
        (10, "沙漠之鹰"),
        (11, ".357左轮"),
        (12, "QSZ92G"),
        (13, "R93")
    ]
    
    for row_idx, name in pistol_rows:
        row = sheet[row_idx]
        market_price = row[2].value if row[2] and row[2].value is not None else 0
        readiness_value = row[3].value if row[3] and row[3].value is not None else 0
        
        # 跳过市场价值为0的手枪（如果启用过滤）
        if filter_zero_price and market_price == 0:
            continue
            
        # 检查是否市场价值≥战备价值+2000
        if filter_overpriced and market_price >= readiness_value + 2000:
            overpriced_weapons.append({
                "category": "手枪",
                "name": name,
                "market_price": market_price,
                "readiness_value": readiness_value
            })
            continue  # 跳过这个武器
        
        weapons["手枪"].append({
            "name": name,
            "market_price": market_price,
            "readiness_value": readiness_value
        })
    
    return weapons, overpriced_weapons

def filter_items_by_target_value(items, target_value, item_type):
    """过滤掉市场价值≥目标战备值的物品"""
    filtered_items = []
    removed_items = []
    
    for item in items:
        if item["market_price"] >= target_value:
            removed_items.append({
                "type": item_type,
                "name": item["name"],
                "market_price": item["market_price"],
                "readiness_value": item["readiness_value"]
            })
        else:
            filtered_items.append(item)
    
    return filtered_items, removed_items

def generate_combinations(target_value, equipment_data, weapon_data, specified_chest=None, specified_backpack=None):
    """生成所有装备组合"""
    start_time = time.time()
    
    # 准备槽位数据（包括空选项）
    none_option = {"name": "无", "market_price": 0, "readiness_value": 0}
    
    # 更新槽位：移除武器槽2
    slots = {
        "weapon1": [none_option],  # 主武器槽
        "pistol": [none_option],   # 手枪槽
        "helmet": [none_option],   # 头盔槽
        "armor": [none_option],    # 护甲槽
        "chest": [none_option],    # 弹挂槽
        "backpack": [none_option]  # 背包槽
    }
    
    # 填充武器槽（非手枪武器）
    for category in weapon_data:
        if category != "手枪":
            for weapon in weapon_data[category]:
                slots["weapon1"].append(weapon)
    
    # 填充手枪槽
    if "手枪" in weapon_data:
        slots["pistol"].extend(weapon_data["手枪"])
    
    # 填充装备槽
    slots["helmet"].extend(equipment_data["头盔"])
    slots["armor"].extend(equipment_data["护甲"])
    
    # 特殊处理：用户指定装备
    if specified_chest:
        slots["chest"] = [specified_chest]  # 只使用指定的胸挂
    else:
        slots["chest"].extend(equipment_data["胸挂"])
    
    if specified_backpack:
        slots["backpack"] = [specified_backpack]  # 只使用指定的背包
    else:
        slots["backpack"].extend(equipment_data["背包"])
    
    # 生成所有组合
    valid_combinations = []
    counter = 0  # 用于生成唯一标识符
    
    for w1 in slots["weapon1"]:
        for pistol in slots["pistol"]:
            for helmet in slots["helmet"]:
                for armor in slots["armor"]:
                    for chest in slots["chest"]:
                        for backpack in slots["backpack"]:
                            total_value = (
                                w1["readiness_value"] +
                                pistol["readiness_value"] +
                                helmet["readiness_value"] +
                                armor["readiness_value"] +
                                chest["readiness_value"] +
                                backpack["readiness_value"]
                            )
                            
                            if total_value >= target_value:
                                total_cost = (
                                    w1["market_price"] +
                                    pistol["market_price"] +
                                    helmet["market_price"] +
                                    armor["market_price"] +
                                    chest["market_price"] +
                                    backpack["market_price"]
                                )
                                
                                counter += 1
                                combination = {
                                    "cost": total_cost,
                                    "value": total_value,
                                    "items": {
                                        "武器槽": w1,
                                        "手枪槽": pistol,
                                        "头盔槽": helmet,
                                        "护甲槽": armor,
                                        "弹挂槽": chest,
                                        "背包槽": backpack
                                    },
                                    "id": counter  # 添加唯一标识符
                                }
                                
                                # 使用堆维护最小成本的3个组合
                                # 添加唯一标识符避免字典比较
                                if len(valid_combinations) < 3:
                                    heapq.heappush(valid_combinations, (-combination["cost"], combination["id"], combination))
                                else:
                                    heapq.heappushpop(valid_combinations, (-combination["cost"], combination["id"], combination))
    
    # 按成本升序排序
    valid_combinations = sorted([c[2] for c in valid_combinations], key=lambda x: x["cost"])
    
    elapsed_time = time.time() - start_time
    return valid_combinations, elapsed_time

def main():
    # 加载基础数据（过滤高价物品和0价值物品）
    equipment_data, overpriced_equipment = load_equipment_data("S5装备价格.xlsx", filter_overpriced=True, filter_zero_price=True)
    weapon_data, overpriced_weapons = load_weapon_data("S5武器价格.xlsx", filter_overpriced=True, filter_zero_price=True)
    
    # 加载完整装备数据（不过滤高价物品和0价值物品）
    raw_equipment_data, _ = load_equipment_data("S5装备价格.xlsx", filter_overpriced=False, filter_zero_price=False)
    
    # 显示被排除的物品
    if overpriced_equipment or overpriced_weapons:
        print("以下物品因市场价值≥战备价值+2000而被排除:")
        
        for item in overpriced_equipment:
            print(f"  [{item['category']}] {item['name']}: 市场价值={item['market_price']}, 战备价值={item['readiness_value']}")
        
        for item in overpriced_weapons:
            print(f"  [{item['category']}] {item['name']}: 市场价值={item['market_price']}, 战备价值={item['readiness_value']}")
        
        print("\n" + "-" * 50 + "\n")
    
    # 用户选择是否指定携行具
    specify_gear = input("是否指定携行具？(是/否): ").strip().lower() in ["是", "y", "yes"]
    specified_chest = None
    specified_backpack = None
    
    if specify_gear:
        # 选择胸挂品质（循环直到选择有效品质或退出）
        while True:
            chest_quality = int(input("选择胸挂品质(0-6, 0=不指定): "))
            if chest_quality == 0:
                print("未指定胸挂，将使用常规选项\n")
                break
                
            # 获取该品质所有胸挂（包含市场价值为0的装备）
            chest_options = [item for item in raw_equipment_data["胸挂"] 
                            if item.get("quality", 0) == chest_quality]
            
            if not chest_options:
                print(f"没有找到{chest_quality}品质的胸挂，请重新选择品质")
                continue
                
            # 显示可选胸挂
            print("\n可选胸挂:")
            for idx, item in enumerate(chest_options, 1):
                # 将市场价值为0的装备标注为"不可交易"
                price_info = "不可交易" if item["market_price"] == 0 else f"市场价:{item['market_price']}"
                print(f"{idx}. {item['name']} ({price_info}, 战备值:{item['readiness_value']})")
            
            # 用户选择具体胸挂
            choice = int(input("选择胸挂编号: "))
            specified_chest = chest_options[choice - 1]
            print(f"已选择胸挂: {specified_chest['name']}\n")
            break
        
        # 选择背包品质（循环直到选择有效品质或退出）
        while True:
            backpack_quality = int(input("选择背包品质(0-6, 0=不指定): "))
            if backpack_quality == 0:
                print("未指定背包，将使用常规选项\n")
                break
                
            backpack_options = [item for item in raw_equipment_data["背包"] 
                              if item.get("quality", 0) == backpack_quality]
            
            if not backpack_options:
                print(f"没有找到{backpack_quality}品质的背包，请重新选择品质")
                continue
                
            print("\n可选背包:")
            for idx, item in enumerate(backpack_options, 1):
                # 将市场价值为0的装备标注为"不可交易"
                price_info = "不可交易" if item["market_price"] == 0 else f"市场价:{item['market_price']}"
                print(f"{idx}. {item['name']} ({price_info}, 战备值:{item['readiness_value']})")
            
            choice = int(input("选择背包编号: "))
            specified_backpack = backpack_options[choice - 1]
            print(f"已选择背包: {specified_backpack['name']}\n")
            break
    
    # 获取目标战备值
    target_value = int(input("请输入目标战备价值: "))
    
    # 保存用户指定的装备
    saved_chest = specified_chest
    saved_backpack = specified_backpack
    
    # 过滤掉市场价值≥目标战备值的物品（跳过用户指定部位）
    removed_by_target = []
    
    # 过滤装备数据（跳过用户指定部位）
    for category in equipment_data:
        # 跳过用户指定部位
        if (category == "胸挂" and saved_chest) or (category == "背包" and saved_backpack):
            continue
            
        filtered, removed = filter_items_by_target_value(equipment_data[category], target_value, category)
        equipment_data[category] = filtered
        removed_by_target.extend(removed)
    
    # 过滤武器数据
    for category in weapon_data:
        filtered, removed = filter_items_by_target_value(weapon_data[category], target_value, category)
        weapon_data[category] = filtered
        removed_by_target.extend(removed)
    
    # 显示因新限制被排除的物品
    if removed_by_target:
        print(f"\n以下物品因市场价值≥目标战备值({target_value})而被排除:")
        for item in removed_by_target:
            print(f"  [{item['type']}] {item['name']}: 市场价值={item['market_price']}, 战备价值={item['readiness_value']}")
        print("\n" + "-" * 50 + "\n")
    
    # 计算组合 - 传递用户指定的装备
    combinations, elapsed_time = generate_combinations(
        target_value, 
        equipment_data, 
        weapon_data,
        specified_chest=saved_chest,
        specified_backpack=saved_backpack
    )
    
    # 输出结果
    print(f"\n计算完成，用时: {elapsed_time:.2f}秒")
    
    if not combinations:
        print("未找到满足条件的组合")
    else:
        print(f"找到 {len(combinations)} 个满足条件的组合，市场价值最低的3个方案:")
        
        for i, combo in enumerate(combinations[:3], 1):
            print(f"\n方案 #{i} - 总市场价值: {combo['cost']}, 总战备价值: {combo['value']}")
            
            for slot, item in combo["items"].items():
                # 当物品为"无"时，只显示名称
                if item["name"] == "无":
                    print(f"  {slot}: {item['name']}")
                else:
                    # 将市场价值为0的装备标注为"不可交易"
                    price_info = "不可交易" if item["market_price"] == 0 else f"市场价值={item['market_price']}"
                    print(f"  {slot}: {item['name']}（{price_info}, 战备价值={item['readiness_value']}）")
    
    input("\n按Enter键退出程序...")

if __name__ == "__main__":
    main()
